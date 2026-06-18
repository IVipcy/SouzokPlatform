# -*- coding: utf-8 -*-
"""
封筒（角２・長形３号 白/茶）テンプレ分割スクリプト（一回限り）

宛名（郵便番号・住所・氏名）を流し込む封筒様式を public/templates/envelope/<key>.xlsx に切り出す。
数式（郵便番号MID・入力元参照・外字）・コメント・枠外・画像を除去。実値は生成時に流し込む。
"""
import openpyxl
from openpyxl.styles import Border, PatternFill
import os

NO_BORDER = Border()
NO_FILL = PatternFill(fill_type=None)

SRC = os.path.join('docs', '作成書類一覧（契約書・委任状・請求書・領収書）.xlsx')
OUT_DIR = os.path.join('public', 'templates', 'envelope')

# idx -> (key, 枠外閾値col)
VARIANTS = {
    33: ('kaku2', 40),        # 角２封筒（外字 AN=40 以降が枠外）
    34: ('naga3_white', 20),  # 長形３号（白）（外字 T=20 以降）
    35: ('naga3_brown', 20),  # 長形３号（茶）
}

os.makedirs(OUT_DIR, exist_ok=True)

for idx, (key, off_col) in VARIANTS.items():
    wb = openpyxl.load_workbook(SRC)
    target = wb.worksheets[idx]
    for s in list(wb.worksheets):
        if s is not target:
            wb.remove(s)
    ws = wb.worksheets[0]
    ws.sheet_state = 'visible'
    ws._images = []
    ws._charts = []

    for mr in list(ws.merged_cells.ranges):
        if mr.min_col >= off_col:
            ws.unmerge_cells(str(mr))
    for row in ws.iter_rows():
        for c in row:
            if c.column >= off_col:
                c.value = None
                c.border = NO_BORDER
                c.fill = NO_FILL
            elif isinstance(c.value, str) and c.value.startswith('='):
                c.value = None
            if c.comment is not None:
                c.comment = None

    out = os.path.join(OUT_DIR, f'{key}.xlsx')
    wb.save(out)
    print(f'  wrote {out}')

print('done', len(VARIANTS), 'files')
