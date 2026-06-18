# -*- coding: utf-8 -*-
"""
確定請求書＋立替実費明細 テンプレ分割スクリプト（一回限り）

確定請求書(行/司)と立替実費明細(行/司)を「1ファイル2シート」にまとめて
public/templates/kakutei/<key>.xlsx に出力する（確定請求が1枚目、立替明細が2枚目）。
数式（入力元参照・合計計算・他シート参照）・コメント・枠外・社印画像を除去。実値は生成時に流し込む。
"""
import openpyxl
from openpyxl.styles import Border, PatternFill
import os

NO_BORDER = Border()
NO_FILL = PatternFill(fill_type=None)

SRC = os.path.join('docs', '作成書類一覧（契約書・委任状・請求書・領収書）.xlsx')
OUT_DIR = os.path.join('public', 'templates', 'kakutei')

# key -> (確定請求シートidx, 立替実費シートidx)
VARIANTS = {
    'kakutei_gyosei': (27, 25),
    'kakutei_shiho': (28, 26),
}


def clean_sheet(ws, off_col, drop_images):
    if drop_images:
        ws._images = []
        ws._charts = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_col >= off_col:
            ws.unmerge_cells(str(mr))
    for row in ws.iter_rows():
        for c in row:
            if c.column >= off_col:
                c.value = None
                c.border = NO_BORDER
                c.fill = NO_FILL
            elif isinstance(c.value, str) and c.value.startswith('='):
                c.value = None
            if c.comment is not None:
                c.comment = None


os.makedirs(OUT_DIR, exist_ok=True)

for key, (ci, ti) in VARIANTS.items():
    wb = openpyxl.load_workbook(SRC)
    kakutei = wb.worksheets[ci]
    tatekae = wb.worksheets[ti]
    for s in list(wb.worksheets):
        if s is not kakutei and s is not tatekae:
            wb.remove(s)
    # 確定請求を1枚目、立替明細を2枚目に並べ替え
    wb._sheets = [kakutei, tatekae]

    clean_sheet(kakutei, off_col=26, drop_images=True)   # 確定: Z以降が枠外(外字AA/経理AG)、社印画像あり
    clean_sheet(tatekae, off_col=11, drop_images=False)  # 立替: K以降が枠外(外字M)

    out = os.path.join(OUT_DIR, f'{key}.xlsx')
    wb.save(out)
    print(f'  wrote {out}  sheets={wb.sheetnames}')

print('done', len(VARIANTS), 'files')
