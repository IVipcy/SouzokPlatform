# -*- coding: utf-8 -*-
"""
請求書・領収書（前受金）テンプレ分割スクリプト（一回限り）

請求書（前受金）行/司・領収書（前受金）行/司を public/templates/invoice/<key>.xlsx に切り出す。
数式（入力元参照・合計計算・領収書の請求書参照）・コメント・枠外(Z=26以降の外字/経理ヘルパー)を除去。
前受金は消費税対象外のため、実行時に合計＝入力額のリテラルを流し込む。
"""
import openpyxl
from openpyxl.styles import Border, PatternFill
import os

NO_BORDER = Border()
NO_FILL = PatternFill(fill_type=None)

SRC = os.path.join('docs', '作成書類一覧（契約書・委任状・請求書・領収書）.xlsx')
OUT_DIR = os.path.join('public', 'templates', 'invoice')
OFF_COL = 26  # Z以降は枠外（外字・経理ヘルパー）。印刷は A〜Y。

VARIANTS = {
    23: 'seikyu_advance_gyosei',
    24: 'seikyu_advance_shiho',
    29: 'ryoshu_advance_gyosei',
    30: 'ryoshu_advance_shiho',
}

os.makedirs(OUT_DIR, exist_ok=True)

for idx, key in VARIANTS.items():
    wb = openpyxl.load_workbook(SRC)
    target = wb.worksheets[idx]
    for s in list(wb.worksheets):
        if s is not target:
            wb.remove(s)
    ws = wb.worksheets[0]

    # テンプレ埋め込み画像（社印）を除去（openpyxl保存で壊れ ExcelJS が読めなくなるため）。
    # 社印は生成時に法人別(gyosei/shiho.png)で再配置する。
    ws._images = []
    ws._charts = []

    # 枠外列のマージ解除
    for mr in list(ws.merged_cells.ranges):
        if mr.min_col >= OFF_COL:
            ws.unmerge_cells(str(mr))

    # 数式・コメント・枠外を除去
    for row in ws.iter_rows():
        for c in row:
            if c.column >= OFF_COL:
                c.value = None
                c.border = NO_BORDER
                c.fill = NO_FILL
            elif isinstance(c.value, str) and c.value.startswith('='):
                c.value = None
            if c.comment is not None:
                c.comment = None

    out = os.path.join(OUT_DIR, f'{key}.xlsx')
    wb.save(out)
    print(f'  wrote {out}')

print('done', len(VARIANTS), 'files')
