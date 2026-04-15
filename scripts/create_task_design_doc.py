"""タスク設計書_v2.xlsx 作成スクリプト"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

wb = Workbook()

# ────────────────────────────────────────────
# ユーティリティ
# ────────────────────────────────────────────
DARK_BLUE  = "1E3A5F"
MID_BLUE   = "2563EB"
LIGHT_BLUE = "DBEAFE"
PALE_BLUE  = "EFF6FF"
PALE_PURPLE= "F5F3FF"
PALE_GREEN = "ECFDF5"
PALE_YELLOW= "FFFBEB"
PALE_ORANGE= "FFF7ED"
PALE_RED   = "FEF2F2"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F3F4F6"
BORDER_COLOR = "D1D5DB"

thin = Side(style="thin", color=BORDER_COLOR)
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, value, bg=DARK_BLUE, fg=WHITE, bold=True, size=11, wrap=False, merge_to_col=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)
    cell.border = thin_border
    if merge_to_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to_col)
    return cell

def cell(ws, row, col, value, bg=WHITE, fg="000000", bold=False, wrap=False, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, color=fg, size=10)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    c.border = thin_border
    return c

def row_cells(ws, row, values, bg=WHITE, wrap=False):
    for col, val in enumerate(values, 1):
        cell(ws, row, col, val, bg=bg, wrap=wrap)

# ────────────────────────────────────────────
# Sheet 1: タスク一括生成ルール
# ────────────────────────────────────────────
ws1 = wb.active
ws1.title = "タスク一括生成ルール"

hdr(ws1, 1, 1, "タスク一括生成 ― 運用ルール", size=14, merge_to_col=4)
ws1.row_dimensions[1].height = 28

hdr(ws1, 3, 1, "生成タイミング・条件", bg=MID_BLUE, merge_to_col=4)
hdr(ws1, 4, 1, "項目", bg=LIGHT_BLUE, fg=DARK_BLUE)
hdr(ws1, 4, 2, "内容", bg=LIGHT_BLUE, fg=DARK_BLUE, merge_to_col=4)

rules = [
    ("生成タイミング",       "案件を「受注」ステータスにしたあと、管理担当が案件詳細画面から「タスク一括生成」ボタンをクリックした時"),
    ("操作権限",             "管理担当（manager）のみ実行可能。アシスタント・受注担当は不可。"),
    ("生成単位",             "案件ごと。同じtemplate_keyのタスクが既に存在する場合はスキップ（重複生成なし）。"),
    ("選択方式",             "フェーズごとにチェックボックスで個別選択 / フェーズ一括選択 / 全選択。不要なタスクは外してから生成する。"),
    ("依存関係の自動作成",   "生成と同時に、定義済みフロールールに基づきタスク間の依存関係（task_dependencies）が自動作成される。"),
    ("追加生成",             "後からでも追加生成可能。生成済みタスクは「生成済」バッジで表示され選択不可となる。"),
    ("手動追加",             "テンプレートにない臨時タスクは「タスクを追加」ボタンから個別作成できる（タイトル・フェーズ・期限・優先度を設定）。"),
    ("依存関係の編集",       "依存関係はテンプレートのフロールールに基づき自動生成のみ。手動編集UIは提供しない（運用簡素化のため）。"),
    ("残高証明請求の特殊仕様", "「残高証明請求」タスクは1タスクで全金融機関をまとめて管理する。タスク生成時に案件の預貯金情報を自動取得。銀行ごとに請求日・到着日を入力し、全銀行の到着日が入力されたらタスクを「完了」にする。"),
]
for i, (k, v) in enumerate(rules):
    bg = PALE_BLUE if i % 2 == 0 else WHITE
    cell(ws1, 5 + i, 1, k, bg=bg, bold=True)
    c = ws1.cell(row=5 + i, column=2, value=v)
    c.font = Font(name="Arial", size=10)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border = thin_border
    ws1.merge_cells(start_row=5 + i, start_column=2, end_row=5 + i, end_column=4)
    ws1.row_dimensions[5 + i].height = 28

hdr(ws1, 13, 1, "フェーズ別 生成推奨タイミング", bg=MID_BLUE, merge_to_col=4)
hdr(ws1, 14, 1, "フェーズ",             bg=LIGHT_BLUE, fg=DARK_BLUE)
hdr(ws1, 14, 2, "推奨生成タイミング",   bg=LIGHT_BLUE, fg=DARK_BLUE, merge_to_col=3)
hdr(ws1, 14, 4, "備考",                 bg=LIGHT_BLUE, fg=DARK_BLUE)

phases = [
    ("Phase1: 相続人調査",    "受注直後（案件対応開始時）",              "最初に必ず生成する"),
    ("Phase2: 財産調査",      "Phase1着手と同時 or 戸籍収集開始後",     "財産の目星がついた時点で"),
    ("Phase3: 不動産・相続税","不動産・相続税が絡む案件のみ",           "不要な場合は生成しない"),
    ("Phase4: 遺産分割",      "財産目録完成後",                         "Phase2完了が前提"),
    ("Phase5: 登記・解約",    "遺産分割協議書が確定後",                 "協議書なしで生成しない"),
    ("Phase6: 完了・精算",    "Phase5の手続きが概ね完了した段階",       "経理担当と連携して実施"),
]
for i, (ph, timing, note) in enumerate(phases):
    bg = PALE_BLUE if i % 2 == 0 else WHITE
    cell(ws1, 15 + i, 1, ph,     bg=bg, bold=True)
    c2 = ws1.cell(row=15 + i, column=2, value=timing)
    c2.font = Font(name="Arial", size=10)
    c2.fill = PatternFill("solid", fgColor=bg)
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c2.border = thin_border
    ws1.merge_cells(start_row=15 + i, start_column=2, end_row=15 + i, end_column=3)
    cell(ws1, 15 + i, 4, note, bg=bg)

ws1.column_dimensions["A"].width = 26
ws1.column_dimensions["B"].width = 40
ws1.column_dimensions["C"].width = 25
ws1.column_dimensions["D"].width = 30

# ────────────────────────────────────────────
# Sheet 2: タスク一覧・完了条件
# ────────────────────────────────────────────
ws2 = wb.create_sheet("タスク一覧・完了条件")

hdr(ws2, 1, 1, "タスク一覧・完了条件・依存関係", size=14, merge_to_col=8)
ws2.row_dimensions[1].height = 28

headers = ["#", "フェーズ", "タスク名", "カテゴリ", "担当ロール", "完了条件", "前提（依存）タスク", "着手可能条件"]
for ci, h in enumerate(headers, 1):
    hdr(ws2, 3, ci, h)

ws2.freeze_panes = "A4"

tasks = [
    # ph, name, cat, role, done_cond, prereq, start_cond, bg
    ("Phase1: 相続人調査","戸籍請求書作成","戸籍","管理担当アシスタント",
     "請求書を作成し管理担当の確認を受けた","（なし）","案件受注後、いつでも着手可",PALE_BLUE),
    ("Phase1: 相続人調査","戸籍郵送手配","戸籍","管理担当アシスタント",
     "請求書・定額小為替・返信用封筒を郵送した","戸籍請求書作成","タスク「戸籍請求書作成」が完了",PALE_BLUE),
    ("Phase1: 相続人調査","戸籍到着確認・読み込み","戸籍","管理担当アシスタント",
     "不足有無を確認し、必要に応じて管理担当に報告した","戸籍郵送手配","タスク「戸籍郵送手配」が完了",PALE_BLUE),
    ("Phase1: 相続人調査","追加戸籍請求","戸籍","管理担当アシスタント",
     "追加分の戸籍を請求した","戸籍到着確認・読み込み","「不足有無」フィールドが「あり（追加請求要）」に入力済",PALE_BLUE),
    ("Phase1: 相続人調査","相続人調査報告書作成","相続人調査","管理担当アシスタント",
     "相続人調査報告書が作成・確認済み","戸籍到着確認・読み込み","「到着日」フィールドが入力済（タスク未完了でも可）",PALE_BLUE),
    ("Phase1: 相続人調査","法定相続情報一覧図作成","相続人調査","管理担当アシスタント",
     "法務局指定書式の一覧図が作成・確認済み","相続人調査報告書作成","タスク「相続人調査報告書作成」が完了",PALE_BLUE),
    ("Phase1: 相続人調査","法定相続情報一覧図 法務局提出","相続人調査","管理担当アシスタント",
     "法務局への申請書類を提出した","法定相続情報一覧図作成","タスク「法定相続情報一覧図作成」が完了",PALE_BLUE),
    ("Phase1: 相続人調査","法定相続情報一覧図 受領","相続人調査","管理担当アシスタント",
     "法務局から一覧図を受領し、必要枚数を確認した","法定相続情報一覧図 法務局提出","タスク「法定相続情報一覧図 法務局提出」が完了",PALE_BLUE),

    ("Phase2: 財産調査","残高証明請求（全銀行まとめて）","金融機関","管理担当アシスタント",
     "全金融機関への請求〜書類到着まで完了。各銀行の「到着日」が全て入力されてから「完了」にする","法定相続情報一覧図 受領","タスク「法定相続情報一覧図 受領」が完了",PALE_PURPLE),
    ("Phase2: 財産調査","証券会社照会","金融機関","管理担当アシスタント",
     "証券会社に照会書類を提出し、残高照会結果を取得した","法定相続情報一覧図 受領","タスク「法定相続情報一覧図 受領」が完了",PALE_PURPLE),
    ("Phase2: 財産調査","保険会社照会","保険","管理担当アシスタント",
     "生命保険協会への照会を申請した","法定相続情報一覧図 受領","タスク「法定相続情報一覧図 受領」が完了",PALE_PURPLE),
    ("Phase2: 財産調査","保険照会結果 到着確認","保険","管理担当アシスタント",
     "照会結果を受領し、保険契約有無を管理担当に報告した","保険会社照会","タスク「保険会社照会」が完了",PALE_PURPLE),
    ("Phase2: 財産調査","年金照会","年金","管理担当アシスタント",
     "年金事務所への照会が完了した","（なし）","案件対応中、いつでも着手可",PALE_PURPLE),
    ("Phase2: 財産調査","不動産調査（謄本・公図等取得）","不動産","管理担当アシスタント",
     "登記情報・公図・地積測量図を取得し、案件レコードに添付した","（なし）","案件対応中、いつでも着手可",PALE_PURPLE),
    ("Phase2: 財産調査","不動産評価額算出","不動産","管理担当アシスタント",
     "不動産の評価額を算出した","不動産調査（謄本・公図等取得）","タスク「不動産調査」が完了",PALE_PURPLE),
    ("Phase2: 財産調査","負債調査（信用情報等）","負債","管理担当アシスタント",
     "信用情報機関への照会が完了した","（なし）","案件対応中、いつでも着手可",PALE_PURPLE),
    ("Phase2: 財産調査","財産目録作成","財産目録","管理担当アシスタント",
     "全財産調査をまとめた財産目録を作成・確認済み",
     "残高証明請求（全銀行）\n証券会社照会\n保険照会結果 到着確認\n不動産評価額算出\n負債調査",
     "上記前提タスクが全て完了",PALE_PURPLE),

    ("Phase3: 不動産・相続税","相続税申告要否判定","相続税","管理担当",
     "相続税申告の要否を判定し、案件詳細に記録した","財産目録作成","タスク「財産目録作成」が完了",PALE_YELLOW),
    ("Phase3: 不動産・相続税","相続税申告書類準備","相続税","管理担当アシスタント",
     "申告に必要な書類を一式準備した","相続税申告要否判定","タスク「相続税申告要否判定」が完了",PALE_YELLOW),
    ("Phase3: 不動産・相続税","税理士への引継ぎ","相続税","管理担当",
     "税理士に案件情報を共有し、連携を開始した","相続税申告書類準備","タスク「相続税申告書類準備」が完了",PALE_YELLOW),
    ("Phase3: 不動産・相続税","不動産鑑定手配","不動産","管理担当",
     "不動産業者に査定依頼し、査定結果を受け取った","不動産調査（謄本・公図等取得）","タスク「不動産調査」が完了",PALE_YELLOW),
    ("Phase3: 不動産・相続税","不動産売却サポート","不動産","管理担当",
     "不動産売却に向けたサポートが完了した","不動産調査（謄本・公図等取得）","タスク「不動産調査」が完了",PALE_YELLOW),

    ("Phase4: 遺産分割","遺産分割協議書 原案作成","遺産分割","管理担当アシスタント",
     "「相続の力」で協議書の原案を作成・確認済み",
     "財産目録作成\n相続税申告要否判定",
     "上記前提タスクが全て完了",PALE_GREEN),
    ("Phase4: 遺産分割","分割案 ご説明","遺産分割","管理担当",
     "依頼者に分割案を説明し、合意を得た","遺産分割協議書 原案作成","タスク「遺産分割協議書 原案作成」が完了",PALE_GREEN),
    ("Phase4: 遺産分割","遺産分割協議書 最終版作成","遺産分割","管理担当アシスタント",
     "協議書の最終版を作成・確認済み","分割案 ご説明","タスク「分割案 ご説明」が完了",PALE_GREEN),
    ("Phase4: 遺産分割","遺産分割協議書 署名捺印手配","遺産分割","管理担当アシスタント",
     "全相続人に郵送して捺印依頼した","遺産分割協議書 最終版作成","タスク「遺産分割協議書 最終版作成」が完了",PALE_GREEN),
    ("Phase4: 遺産分割","遺産分割協議書 回収確認","遺産分割","管理担当アシスタント",
     "全相続人の署名捺印済み協議書を回収・照合した","遺産分割協議書 署名捺印手配","タスク「遺産分割協議書 署名捺印手配」が完了",PALE_GREEN),

    ("Phase5: 登記・解約","登記申請書類作成","登記","管理担当アシスタント",
     "登記申請書類一式を作成・確認済み","遺産分割協議書 回収確認","タスク「遺産分割協議書 回収確認」が完了",PALE_ORANGE),
    ("Phase5: 登記・解約","登記申請（法務局）","登記","管理担当アシスタント",
     "法務局に申請書類を提出した","登記申請書類作成","タスク「登記申請書類作成」が完了",PALE_ORANGE),
    ("Phase5: 登記・解約","登記完了確認・謄本取得","登記","管理担当アシスタント",
     "登記完了を確認し、謄本を取得した","登記申請（法務局）","タスク「登記申請（法務局）」が完了",PALE_ORANGE),
    ("Phase5: 登記・解約","預貯金解約・名義変更手続き","金融機関","管理担当アシスタント",
     "口座解約・名義変更が完了し、入金を確認した","遺産分割協議書 回収確認","タスク「遺産分割協議書 回収確認」が完了",PALE_ORANGE),
    ("Phase5: 登記・解約","証券口座移管・解約手続き","金融機関","管理担当アシスタント",
     "証券口座の移管または解約が完了した","遺産分割協議書 回収確認","タスク「遺産分割協議書 回収確認」が完了",PALE_ORANGE),
    ("Phase5: 登記・解約","保険金請求手続き","保険","管理担当アシスタント",
     "保険金の請求手続きが完了した","遺産分割協議書 回収確認","タスク「遺産分割協議書 回収確認」が完了",PALE_ORANGE),
    ("Phase5: 登記・解約","自動車名義変更","その他","管理担当アシスタント",
     "自動車の名義変更が完了した","遺産分割協議書 回収確認","タスク「遺産分割協議書 回収確認」が完了",PALE_ORANGE),

    ("Phase6: 完了・精算","分配金計算書作成","精算","経理担当",
     "立替実費を集計し、計算書を作成した",
     "登記完了確認・謄本取得\n預貯金解約・名義変更手続き\n証券口座移管・解約手続き\n保険金請求手続き",
     "上記前提タスクが全て完了",PALE_RED),
    ("Phase6: 完了・精算","報酬請求書作成","精算","経理担当",
     "依頼者宛の請求書を作成し、受注担当に確認を取った","分配金計算書作成","タスク「分配金計算書作成」が完了",PALE_RED),
    ("Phase6: 完了・精算","入金確認","精算","経理担当",
     "依頼者から入金が確認できた","報酬請求書作成","タスク「報酬請求書作成」が完了",PALE_RED),
    ("Phase6: 完了・精算","分配金送金実行","精算","経理担当",
     "各相続人への分配金を送金した","入金確認","タスク「入金確認」が完了",PALE_RED),
    ("Phase6: 完了・精算","納品書類一式作成","納品","管理担当アシスタント",
     "原本書類を依頼者に発送し、送付書を作成した","分配金送金実行","タスク「分配金送金実行」が完了",PALE_RED),
    ("Phase6: 完了・精算","案件クローズ処理","納品","管理担当",
     "案件ステータスを「完了」に変更し、クローズ処理を行った","納品書類一式作成","タスク「納品書類一式作成」が完了",PALE_RED),
]

for i, (ph, name, cat, role, done, prereq, start, bg) in enumerate(tasks, 1):
    r = 3 + i
    cell(ws2, r, 1, i,      bg=bg, align="center")
    cell(ws2, r, 2, ph,     bg=bg, bold=True)
    cell(ws2, r, 3, name,   bg=bg, bold=True)
    cell(ws2, r, 4, cat,    bg=bg)
    cell(ws2, r, 5, role,   bg=bg)
    cell(ws2, r, 6, done,   bg=bg, wrap=True)
    cell(ws2, r, 7, prereq, bg=bg, wrap=True)
    cell(ws2, r, 8, start,  bg=bg, wrap=True)
    ws2.row_dimensions[r].height = 45

ws2.column_dimensions["A"].width = 4
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 28
ws2.column_dimensions["D"].width = 14
ws2.column_dimensions["E"].width = 22
ws2.column_dimensions["F"].width = 40
ws2.column_dimensions["G"].width = 40
ws2.column_dimensions["H"].width = 38

# ────────────────────────────────────────────
# Sheet 3: タスク別 入力項目
# ────────────────────────────────────────────
ws3 = wb.create_sheet("タスク別入力項目")
hdr(ws3, 1, 1, "タスク別 カテゴリ入力項目（ext_data）", size=14, merge_to_col=5)
ws3.row_dimensions[1].height = 28

note = ws3.cell(row=3, column=1,
    value="※ 各タスクのカテゴリに応じて、タスク詳細画面の最下部「📝 作業内容」セクションに専用の入力フォームが表示されます（カテゴリ別にセクション名は変えず「作業内容」で統一）。データはext_data（JSONB）に保存されます。")
note.font = Font(name="Arial", italic=True, color="6B7280", size=9)
note.alignment = Alignment(wrap_text=True)
ws3.merge_cells("A3:E3")
ws3.row_dimensions[3].height = 20

for ci, h in enumerate(["カテゴリ","フィールドキー","項目名","種別","説明・選択肢"], 1):
    hdr(ws3, 5, ci, h)
ws3.freeze_panes = "A6"

fields = [
    # category, key, label, type, note, bg
    ("戸籍","city","請求先市区町村","テキスト","",PALE_BLUE),
    ("戸籍","kosekiType","請求した戸籍の種類","プルダウン","全部事項証明 / 除籍謄本 / 改製原戸籍 / 戸籍の附票 / 住民票",PALE_BLUE),
    ("戸籍","reqDate","請求日","日付","",PALE_BLUE),
    ("戸籍","arrDate","到着日","日付","★ チェックポイント: 入力済みで「相続人調査報告書作成」の着手が可能になる",PALE_BLUE),
    ("戸籍","shortage","不足有無","プルダウン","なし / あり（追加請求要） ★「あり」で「追加戸籍請求」が着手可能に",PALE_BLUE),
    ("戸籍","memo","メモ","テキストエリア","",PALE_BLUE),

    # 残高証明請求タスク（bank_balance_request）は専用の複数銀行UIを使用
    ("残高証明請求\n【専用UI】","ext_data.banks\n（配列）","金融機関リスト","JSON配列","★ 案件の「財産情報（預貯金）」から自動取得。タスク一括生成時に埋め込まれる。",PALE_PURPLE),
    ("残高証明請求\n【銀行ごとに入力】","banks[].institution_name","金融機関名","テキスト（表示のみ）","案件の財産情報から自動取得。編集不可。",PALE_PURPLE),
    ("残高証明請求\n【銀行ごとに入力】","banks[].branch_name","支店名","テキスト（表示のみ）","案件の財産情報から自動取得。編集不可。",PALE_PURPLE),
    ("残高証明請求\n【銀行ごとに入力】","banks[].frozen","凍結済","チェックボックス","凍結依頼を同時に実施した場合にチェック",PALE_PURPLE),
    ("残高証明請求\n【銀行ごとに入力】","banks[].reqDate","請求日","日付","各銀行に郵送した日付を入力",PALE_PURPLE),
    ("残高証明請求\n【銀行ごとに入力】","banks[].arrDate","到着日","日付","★ 書類が届いたら入力。全銀行に入力されたらタスクを「完了」にする",PALE_PURPLE),

    # その他の財産調査(預貯金)カテゴリタスク用（個別銀行タスクの場合）
    ("財産調査(預貯金)\n【個別タスク用】","bank","金融機関名","テキスト","",PALE_PURPLE),
    ("財産調査(預貯金)\n【個別タスク用】","branch","支店","テキスト","",PALE_PURPLE),
    ("財産調査(預貯金)\n【個別タスク用】","investigationType","調査種別","プルダウン","現存確認 / 残高証明 / 残高証明(相続開始日) / 取引履歴 / 経過利息 / 全店調査",PALE_PURPLE),
    ("財産調査(預貯金)\n【個別タスク用】","frozen","凍結済","チェックボックス","",PALE_PURPLE),
    ("財産調査(預貯金)\n【個別タスク用】","reqDate","郵送・窓口請求日","日付","",PALE_PURPLE),
    ("財産調査(預貯金)\n【個別タスク用】","arrDate","書類到着日（完了日）","日付","",PALE_PURPLE),
    ("財産調査(預貯金)\n【個別タスク用】","memo","メモ","テキストエリア","",PALE_PURPLE),

    ("財産調査(証券)","company","証券会社名/信託銀行名","テキスト","",PALE_GREEN),
    ("財産調査(証券)","branch","支店","テキスト","",PALE_GREEN),
    ("財産調査(証券)","stock","銘柄名","テキスト","",PALE_GREEN),
    ("財産調査(証券)","investigationType","調査種別","プルダウン","残高照会 / 所有株式数証明 / 未受領配当金 / 配当金支払通知書 / ほふり照会",PALE_GREEN),
    ("財産調査(証券)","reqDate","郵送・窓口請求日","日付","",PALE_GREEN),
    ("財産調査(証券)","arrDate","書類到着日（完了日）","日付","",PALE_GREEN),
    ("財産調査(証券)","memo","メモ","テキストエリア","",PALE_GREEN),

    ("解約手続き","bank","金融機関名/証券会社名","テキスト","",PALE_YELLOW),
    ("解約手続き","accountType","口座種別","プルダウン","預貯金 / 証券 / 信託銀行",PALE_YELLOW),
    ("解約手続き","checkSheet","確認シートチェック","チェックボックス","",PALE_YELLOW),
    ("解約手続き","transferTo","解約金振込先","テキスト","",PALE_YELLOW),
    ("解約手続き","processDate","手続日","日付","",PALE_YELLOW),
    ("解約手続き","completeDate","完了日","日付","",PALE_YELLOW),
    ("解約手続き","memo","メモ","テキストエリア","",PALE_YELLOW),

    ("不動産/登記申請書作成","addr","不動産所在地","テキスト","",PALE_ORANGE),
    ("不動産/登記申請書作成","propType","物件種別","プルダウン","土地 / 建物 / マンション / その他",PALE_ORANGE),
    ("不動産/登記申請書作成","agent","査定依頼先","テキスト","",PALE_ORANGE),
    ("不動産/登記申請書作成","agentReqDate","査定依頼日","日付","",PALE_ORANGE),
    ("不動産/登記申請書作成","amount","査定金額","通貨","",PALE_ORANGE),
    ("不動産/登記申請書作成","applyDate","登記申請日","日付","",PALE_ORANGE),
    ("不動産/登記申請書作成","completeDate","登記完了日","日付","",PALE_ORANGE),
    ("不動産/登記申請書作成","memo","メモ","テキストエリア","",PALE_ORANGE),

    ("税理士連携","taxAdvisor","税理士名","テキスト","",PALE_RED),
    ("税理士連携","contactDate","連絡日","日付","",PALE_RED),
    ("税理士連携","instruction","指示内容","テキストエリア","",PALE_RED),
    ("税理士連携","deadline","申告期限","日付","",PALE_RED),
    ("税理士連携","memo","メモ","テキストエリア","",PALE_RED),

    ("協議書/財産目録","sendMethod","送付方法","プルダウン","郵送 / メール / 対面 / 依頼者から各相続人へ / OCから各相続人へ",PALE_BLUE),
    ("協議書/財産目録","sendDate","送付日","日付","",PALE_BLUE),
    ("協議書/財産目録","replyDue","回答期限","日付","",PALE_BLUE),
    ("協議書/財産目録","allCollected","全員回収完了","チェックボックス","",PALE_BLUE),
    ("協議書/財産目録","memo","メモ","テキストエリア","",PALE_BLUE),

    ("経理/精算","settleStatus","精算ステータス","プルダウン","実費集計中 / 請求書作成中 / 請求済 / 入金待ち / 入金確認済 / 完了",PALE_PURPLE),
    ("経理/精算","deliveryDate","原本納品日","日付","",PALE_PURPLE),
    ("経理/精算","receiptCollected","原本受領書回収","チェックボックス","",PALE_PURPLE),
]

for i, (cat, key, label, ftype, fnote, bg) in enumerate(fields):
    r = 6 + i
    cell(ws3, r, 1, cat,   bg=bg, bold=True)
    cell(ws3, r, 2, key,   bg=bg)
    cell(ws3, r, 3, label, bg=bg)
    cell(ws3, r, 4, ftype, bg=bg)
    cell(ws3, r, 5, fnote, bg=bg, wrap=True)
    if fnote and "★" in fnote:
        ws3.cell(row=r, column=5).font = Font(name="Arial", size=10, bold=True, color="DC2626")

ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 22
ws3.column_dimensions["C"].width = 25
ws3.column_dimensions["D"].width = 16
ws3.column_dimensions["E"].width = 60

# ────────────────────────────────────────────
# Sheet 4: 作業手順テンプレート
# ────────────────────────────────────────────
ws4 = wb.create_sheet("作業手順テンプレート")
hdr(ws4, 1, 1, "作業手順テンプレート（タスク詳細画面「作業手順」欄）", size=14, merge_to_col=3)
ws4.row_dimensions[1].height = 28

note4 = ws4.cell(row=3, column=1,
    value="※ 各タスクの「作業手順」欄に表示されます。□はチェック項目です。")
note4.font = Font(name="Arial", italic=True, color="6B7280", size=9)
ws4.merge_cells("A3:C3")

for ci, h in enumerate(["タスク名","担当ロール","作業手順・ポイント"], 1):
    hdr(ws4, 5, ci, h)
ws4.freeze_panes = "A6"

procedures = [
    ("戸籍請求書作成","管理担当アシスタント",
"""【作業内容】被相続人の出生〜死亡をたどる戸籍の請求書を作成します。

【手順】
□ 案件詳細の「戸籍請求関連」セクションを開く
□ 「請求書パターン」「筆頭者」「使用目的」「請求理由」を確認する
□ 共有フォルダのテンプレートを開き、案件情報を転記する
□ 特記事項がある場合は請求書に反映する
□ 作成後、管理担当に確認を依頼してからステータスを「完了」にする

【ポイント】
・外字がある場合は必ず確認。手書きまたはPDF添付で対応。
・筆頭者が被相続人と異なる場合がある（婚姻前の本籍地など）。"""),

    ("戸籍郵送手配","管理担当アシスタント",
"""【作業内容】作成した戸籍請求書を各市区町村役所へ郵送します。

【手順】
□ 請求書・定額小為替・返信用封筒をセットで封入する
□ 宛先（各市区町村の戸籍担当窓口）を確認する
□ 郵送後、コメントに「〇月〇日郵送済み」と記録する
□ ステータスを「対応中（到着待ち）」に変更する

【ポイント】
・定額小為替の金額は市区町村ごとに異なる（通常450円/通）。
・返信用封筒には差出人住所を記載すること。"""),

    ("戸籍到着確認・読み込み","管理担当アシスタント",
"""【作業内容】届いた戸籍書類が揃っているか確認します。

【手順】
□ 届いた戸籍に不足がないか確認する（出生〜死亡の連続性）
□ 相続人が全員確認できるかチェックする
□ 不足がある場合は管理担当に報告し、追加請求の指示を仰ぐ
□ 揃っていればステータスを「完了」にする

【ポイント】
・転籍を繰り返している場合、複数の市区町村から取得が必要。
・不足があっても勝手に追加請求しない。必ず管理担当に確認。"""),

    ("法定相続情報一覧図作成","管理担当アシスタント",
"""【作業内容】法務局に提出する「法定相続情報一覧図」を作成します。

【手順】
□ 相続関係図をもとに、法務局指定の書式で一覧図を作成する
□ 申出人（依頼者）の住所・氏名を正確に記載する
□ 管理担当の確認を受けてから法務局へ提出する
□ 交付された一覧図の枚数を確認し、案件レコードに添付する

【ポイント】
・金融機関への相続手続き時に必要な枚数分を取得すること。
・通常、金融機関の数＋α枚取得しておくとスムーズ。"""),

    ("残高証明請求（全銀行まとめて）","管理担当アシスタント",
"""【作業内容】全金融機関に残高証明書を請求し、書類が揃うまで管理します。
タスク詳細に金融機関ごとの進捗表が表示されます（「残高証明請求（預貯金）」セクション）。

【手順】
□ タスク詳細の金融機関リストを確認する（案件の財産情報から自動取得）
□ 各金融機関の相続センターに電話し「相続手続きの書類を送付してほしい」と依頼する
□ 書類が届いたら残高証明請求書・委任状・印鑑証明書・法定相続情報を準備する
□ 原本一式と一緒に郵送し、金融機関ごとに「請求日」を入力する
□ 書類が届いたら各金融機関の「到着日」を入力する
□ 全金融機関の「到着日」が入力できたらタスクを「完了」にする

【ポイント】
・口座が凍結されていない場合は凍結依頼も同時に行い、「凍結済」にチェックを入れる。
・「相続税申告あり」の案件は相続開始日時点の残高も必要（管理担当に確認）。
・口座が凍結されていない場合は凍結依頼も同時に行う。"""),

    ("証券会社照会","管理担当アシスタント",
"""【作業内容】証券会社に残高照会を行います。

【手順】
□ 対象の証券会社・支店名を確認する
□ 相続手続き窓口に連絡し、残高照会書類の送付を依頼する
□ 書類が届いたら準備して郵送する
□ 残高照会結果が届いたら案件レコードに添付してステータスを「完了」にする

【ポイント】
・証券会社によって手続き窓口（支店 or 相続センター）が異なる。
・手数料が発生する場合は管理担当に確認してから振り込む。"""),

    ("保険会社照会","管理担当アシスタント",
"""【作業内容】生命保険協会に被相続人の保険契約を照会します。

【手順】
□ 生命保険協会の照会申請書を準備する
□ 必要書類（法定相続情報・申請者の身分証明等）と一緒に送付する
□ 結果が届いたら（通常2〜3週間）案件レコードに添付する
□ 保険契約が判明した場合は管理担当に報告する"""),

    ("不動産調査（謄本・公図等取得）","管理担当アシスタント",
"""【作業内容】登記簿図書館を使って不動産の登記情報・公図・地積測量図を取得します。

【手順】
□ 案件の不動産明細から地番を確認する
□ 登記簿図書館にログインし、地番を入力して取得する
□ 登記情報・公図・地積測量図をダウンロードする
□ 案件レコードに添付してステータスを「完了」にする

【ポイント】
・地番（ちばん）は住居表示（住所）と異なる場合があるので注意。"""),

    ("不動産鑑定手配","管理担当",
"""【作業内容】売却予定の不動産の査定を不動産業者に依頼します。

【手順】
□ 管理担当から担当業者の指定があるか確認する
□ 業者に査定依頼の連絡を入れる（電話またはメール）
□ 査定日時を調整し、依頼者に日程を連絡する
□ 査定結果が出たら管理担当に報告してステータスを「完了」にする

【ポイント】
・受注直後に早めに動くこと。物件が複数の場合はタスクが物件ごとに生成される。"""),

    ("税理士への引継ぎ","管理担当",
"""【作業内容】相続税申告が必要な案件で税理士を紹介・連携します。

【手順】
□ 案件詳細の「相続税申告」セクションで申告期限を確認する
□ 管理担当に税理士紹介の指示を仰ぐ
□ 税理士事務所に案件情報を共有する
□ 連携完了後、税理士名を案件詳細に記録してステータスを「完了」にする

【ポイント】
・申告期限は相続開始日から10ヶ月。余裕をもって動くこと。"""),

    ("財産目録作成","管理担当アシスタント",
"""【作業内容】全財産調査の結果をまとめた財産目録を作成します。

【手順】
□ Phase2の全タスクが完了していることを確認する
□ 不動産・金融資産・証券・負債の情報をまとめる
□ 所定のテンプレートに転記して財産目録を作成する
□ 管理担当の確認を受けてからステータスを「完了」にする

【ポイント】
・Phase2が全て完了してから着手すること。
・金額は証明書の数字をそのまま使い、自己判断で計算しない。"""),

    ("遺産分割協議書 原案作成","管理担当アシスタント",
"""【作業内容】「相続の力」を使って遺産分割協議書を作成します。

【手順】
□ 分割方針が確定していることを確認する
□ 「相続の力」に不動産情報を入力してWordファイルを出力する
□ 金融資産の情報を手入力で追記する
□ 管理担当の確認を受けてからステータスを「完了」にする

【ポイント】
・金融資産の記載は手入力で追記必須。内容の判断は管理担当が行う。"""),

    ("遺産分割協議書 署名捺印手配","管理担当アシスタント",
"""【作業内容】相続人全員に遺産分割協議書への署名捺印を依頼し、回収します。

【手順】
□ 協議書を相続人の人数分コピーして郵送準備する
□ 返信用封筒を同封して各相続人に郵送する
□ 返送されてきたら署名・捺印・印鑑の照合をする
□ 全員分揃ったら管理担当に報告してステータスを「完了」にする

【ポイント】
・捺印は実印（印鑑証明書と同じもの）であることを確認。
・押印ミス等があれば勝手に対応せず、管理担当に報告する。"""),

    ("登記申請（法務局）","管理担当アシスタント",
"""【作業内容】協議書に基づき、不動産の名義変更登記を法務局に申請します。

【手順】
□ 協議書の署名捺印回収が完了していることを確認する
□ 管理担当が登記申請書類を作成・確認する（アシスタントは補助）
□ 管轄法務局に申請書類一式を持参または郵送する
□ 完了証が届いたら案件レコードに添付してステータスを「完了」にする

【ポイント】
・同一法務局管轄の不動産は一括申請できる。"""),

    ("預貯金解約・名義変更手続き","管理担当アシスタント",
"""【作業内容】協議書に基づき、預貯金口座の解約手続きを行います。

【手順】
□ 協議書の署名捺印回収が完了していることを確認する
□ 解約に必要な書類（相続届・委任状・印鑑証明・協議書・通帳等）を準備する
□ 郵送または窓口来店で手続きを行う
□ 解約完了・入金確認後、コメントに記録してステータスを「完了」にする

【ポイント】
・通帳・キャッシュカードがある場合は一緒に提出すること。
・端数の振込先は協議書または管理担当の指示を確認すること。"""),

    ("証券口座移管・解約手続き","管理担当アシスタント",
"""【作業内容】証券口座の相続手続き（株式移管または売却・解約）を行います。

【手順】
□ 対象証券会社の相続手続きの方法を確認する
□ 必要書類（相続届・協議書・印鑑証明等）を準備して提出する
□ 手数料が発生する場合は管理担当に確認してから振り込む
□ 手続き完了後、コメントに記録してステータスを「完了」にする

【ポイント】
・証券は解約（売却）か移管かを協議書で確認すること。
・手数料は管理担当の承認を得てから振り込む。"""),

    ("分配金計算書作成","経理担当",
"""【作業内容】案件で立て替えた実費（郵便代・手数料等）を集計します。

【手順】
□ Phase5の全タスクが完了していることを確認する
□ 実費の明細を請求書モジュールに入力する（金額は管理担当が確認）
□ 管理担当に実費明細の確認を依頼する
□ 確認完了後、ステータスを「完了」にする

【ポイント】
・領収書・記録が残っているか確認してから集計すること。"""),

    ("報酬請求書作成","経理担当",
"""【作業内容】依頼者への請求書を作成します。

【手順】
□ 立替実費計算が完了していることを確認する
□ 案件の報酬金額・立替実費・確定金額を確認する
□ 請求書テンプレートに情報を入力して作成する
□ 受注担当に確認を依頼してからステータスを「完了」にする"""),

    ("納品書類一式作成","管理担当アシスタント",
"""【作業内容】戸籍等の原本書類を依頼者に返却・納品します。

【手順】
□ 請求書作成が完了していることを確認する
□ 返却する原本書類（戸籍・印鑑証明等）を一式揃える
□ 送付書（内容物リスト）を作成して同封する
□ レタパックまたは書留で依頼者に発送する
□ 発送後、コメントに「〇月〇日 発送済み」と記録してステータスを「完了」にする

【ポイント】
・送付書は必ず作成し、コピーを案件ファイルに保管する。
・受領確認（受領書の返送）が必要な案件は管理担当に確認する。"""),
]

for i, (name, role, proc) in enumerate(procedures):
    r = 6 + i
    bg = PALE_BLUE if i % 2 == 0 else WHITE
    cell(ws4, r, 1, name, bg=bg, bold=True)
    cell(ws4, r, 2, role, bg=bg)
    c = ws4.cell(row=r, column=3, value=proc)
    c.font = Font(name="Arial", size=9)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border = thin_border
    ws4.row_dimensions[r].height = 130

ws4.column_dimensions["A"].width = 28
ws4.column_dimensions["B"].width = 22
ws4.column_dimensions["C"].width = 72

# ────────────────────────────────────────────
# Sheet 5: 画面構成（タスク詳細・タスク一覧・ダッシュボード）
# ────────────────────────────────────────────
ws5 = wb.create_sheet("画面構成")
hdr(ws5, 1, 1, "画面構成・運用ルール（現場向けUI最適化）", size=14, merge_to_col=3)
ws5.row_dimensions[1].height = 28

note5 = ws5.cell(row=3, column=1,
    value="※ 「複数のパートタイマーが日替わりで着手前タスクをピックアップ → 完了 or 途中で帰る」という運用を前提に最適化された画面構成。")
note5.font = Font(name="Arial", italic=True, color="6B7280", size=9)
note5.alignment = Alignment(wrap_text=True)
ws5.merge_cells("A3:C3")
ws5.row_dimensions[3].height = 28

# 5-1 タスク詳細画面の構成
hdr(ws5, 5, 1, "■ タスク詳細画面の構成（上から順）", bg=MID_BLUE, merge_to_col=3)
hdr(ws5, 6, 1, "セクション", bg=LIGHT_BLUE, fg=DARK_BLUE)
hdr(ws5, 6, 2, "目的・内容", bg=LIGHT_BLUE, fg=DARK_BLUE)
hdr(ws5, 6, 3, "備考", bg=LIGHT_BLUE, fg=DARK_BLUE)

detail_sections = [
    ("ヘッダー",
     "ID / フェーズ / カテゴリ バッジ、タスク名、案件リンク、着手/完了ボタン、ステータスフロー（着手前→対応中→完了）",
     "着手ボタン下に「作業を始める前に押す」、完了ボタン下に「完了条件を満たしたら押す」のヒント表示"),
    ("👉 今やること",
     "✅ 完了条件（このタスクを完了にするタイミング）／📋 作業手順（チェックリスト形式）を最上段に強調表示",
     "完了条件は src/lib/taskCompletionConditions.ts、作業手順は task_templates.procedure_text から取得"),
    ("📝 基本情報",
     "件名・起票日・期限・ステータス・優先度・フェーズ・カテゴリ・備考",
     "旧「内容・分類」セクションを統合。表題（作業内容）フィールドは作業手順と重複するため削除"),
    ("👤 着手者・作業履歴",
     "現在の着手者を緑色カードで表示。下部にこのタスクの活動履歴",
     "未着手の場合は「まだ誰も着手していません」＋ ▶着手するボタン"),
    ("📝 作業内容",
     "カテゴリ別の入力フォーム（請求先市町村・到着日など）",
     "全カテゴリでセクション名を「作業内容」に統一。カテゴリ識別はヘッダーのバッジで行う"),
]

for i, (sec, purpose, note) in enumerate(detail_sections):
    bg = PALE_BLUE if i % 2 == 0 else WHITE
    cell(ws5, 7 + i, 1, sec, bg=bg, bold=True)
    cell(ws5, 7 + i, 2, purpose, bg=bg, wrap=True)
    cell(ws5, 7 + i, 3, note, bg=bg, wrap=True)
    ws5.row_dimensions[7 + i].height = 60

# 5-2 タスク一覧画面の運用
start_row = 7 + len(detail_sections) + 2
hdr(ws5, start_row, 1, "■ タスク一覧画面の運用ルール", bg=MID_BLUE, merge_to_col=3)
hdr(ws5, start_row + 1, 1, "項目", bg=LIGHT_BLUE, fg=DARK_BLUE)
hdr(ws5, start_row + 1, 2, "仕様", bg=LIGHT_BLUE, fg=DARK_BLUE, merge_to_col=3)

list_rules = [
    ("デフォルトフィルター",
     "「着手前」ステータスのみ表示。出勤したパートが即「今日やれるタスク」を選べる。"),
    ("要対応セクション",
     "画面上部に赤枠で固定表示。着手前かつ「期限超過」または「急ぎ」優先度のタスクが自動で最上段に集約。"),
    ("対応中タスクの表示",
     "着手者アバターを通常より大きく（22px）青枠＋太字で強調。誰かがやっている作業を一目で判別、重複着手を防止。"),
    ("ステータスKPIカード",
     "全タスク／着手前／対応中／完了 の4枚。クリックで該当ステータスにフィルター切替。"),
    ("「自分が着手中」フィルター",
     "ボタンで自分が着手or主担当のタスクのみ表示可能。途中で帰った後に再開する用。"),
    ("グループ化",
     "デフォルト：フェーズ別。「📋 案件別」ボタンで案件別グループに切替可能。"),
    ("操作ボタン",
     "▶ 着手する（緑）／ ✅ 完了にする（青）／ ✅ 完了（バッジ）の3状態のみ。着手中→着手前への戻し機能は提供しない。"),
]

for i, (k, v) in enumerate(list_rules):
    bg = PALE_PURPLE if i % 2 == 0 else WHITE
    r = start_row + 2 + i
    cell(ws5, r, 1, k, bg=bg, bold=True)
    c = ws5.cell(row=r, column=2, value=v)
    c.font = Font(name="Arial", size=10)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border = thin_border
    ws5.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws5.row_dimensions[r].height = 38

# 5-3 ダッシュボード
start_row2 = start_row + 2 + len(list_rules) + 2
hdr(ws5, start_row2, 1, "■ ダッシュボード（経営者・管理者向け）", bg=MID_BLUE, merge_to_col=3)
hdr(ws5, start_row2 + 1, 1, "セクション", bg=LIGHT_BLUE, fg=DARK_BLUE)
hdr(ws5, start_row2 + 1, 2, "内容", bg=LIGHT_BLUE, fg=DARK_BLUE, merge_to_col=3)

dashboard_items = [
    ("KPIカード（5枚）", "総案件数 / 対応中タスク / 完了案件 / 検討中案件 / メンバー数。経営者が全体俯瞰するための数値。"),
    ("最近の案件", "最新5件の案件カード。受注担当者アバター付き。"),
    ("期限が近いタスク", "期限指定された未完了タスクを期限昇順で最大8件。担当者アバター付き。"),
    ("マイタスク表示", "提供しない（複数パートが日替わりで動く運用に合わないため削除済み）。現場は /tasks で完結。"),
]

for i, (k, v) in enumerate(dashboard_items):
    bg = PALE_GREEN if i % 2 == 0 else WHITE
    r = start_row2 + 2 + i
    cell(ws5, r, 1, k, bg=bg, bold=True)
    c = ws5.cell(row=r, column=2, value=v)
    c.font = Font(name="Arial", size=10)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border = thin_border
    ws5.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws5.row_dimensions[r].height = 32

# 5-4 InlineEdit UX
start_row3 = start_row2 + 2 + len(dashboard_items) + 2
hdr(ws5, start_row3, 1, "■ インライン編集（InlineEdit）の見た目", bg=MID_BLUE, merge_to_col=3)
hdr(ws5, start_row3 + 1, 1, "状態", bg=LIGHT_BLUE, fg=DARK_BLUE)
hdr(ws5, start_row3 + 1, 2, "見た目", bg=LIGHT_BLUE, fg=DARK_BLUE, merge_to_col=3)

inline_items = [
    ("通常表示", "値の下に点線アンダーライン。クリック可能なことを視覚的に示す。"),
    ("ホバー時", "薄い青背景（hover:bg-blue-50）に変化。tooltip「クリックして入力」を表示。"),
    ("空欄時", "薄いグレーで「クリックして入力」「クリックして選択」「クリックして日付入力」プレースホルダ。"),
    ("編集中", "input/select/textarea に切替。Enter or Blur で保存、Escでキャンセル。"),
]

for i, (k, v) in enumerate(inline_items):
    bg = PALE_YELLOW if i % 2 == 0 else WHITE
    r = start_row3 + 2 + i
    cell(ws5, r, 1, k, bg=bg, bold=True)
    c = ws5.cell(row=r, column=2, value=v)
    c.font = Font(name="Arial", size=10)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c.border = thin_border
    ws5.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws5.row_dimensions[r].height = 28

ws5.column_dimensions["A"].width = 28
ws5.column_dimensions["B"].width = 50
ws5.column_dimensions["C"].width = 45

# ────────────────────────────────────────────
# 保存
# ────────────────────────────────────────────
out = r"C:\Users\sugur\Desktop\相続プラットフォーム\docs\タスク設計書_v2.xlsx"
wb.save(out)
print(f"saved: {out}")
