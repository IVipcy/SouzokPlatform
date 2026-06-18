# -*- coding: utf-8 -*-
"""
委任状テンプレ分割スクリプト（一回限り）

docs/作成書類一覧（契約書・委任状・請求書・領収書）.xlsx の委任状シート（index 7〜16）を
変種ごとに public/templates/ininjo/<key>.xlsx として切り出す。

切り出し時に「お客さんに提出するときは削除する」参照データを除去する:
  - すべての数式セル（=入力元! 参照・=IF/=VLOOKUP のヘルパー）→ 空に（実行時にリテラル流し込み）
  - 印刷枠外の列（Z列以降＝外字ヘルパー・署名参照マスタ等）→ 空に
法定相続情報(7)のみ 代理人住所(G10) が VLOOKUP のため、行政書士法人オーシャンの住所リテラルに解決してから数式を消す。
"""
import openpyxl
from openpyxl.styles import Border, PatternFill
import os

NO_BORDER = Border()
NO_FILL = PatternFill(fill_type=None)

SRC = os.path.join('docs', '作成書類一覧（契約書・委任状・請求書・領収書）.xlsx')
OUT_DIR = os.path.join('public', 'templates', 'ininjo')
GYOSEI_ADDR = '横浜市西区高島２丁目１４－１７ クレアトール横浜ビル５階'

# index -> 出力キー
VARIANTS = {
    7:  'houtei',          # 法定相続情報証明
    8:  'rengo',           # 行・司連名（相続）
    9:  'rengo_yokin',     # 行・司連名（預金解約あり）
    10: 'rengo_touki',     # 行・司連名（登記のみ）
    11: 'gyosei_souzoku',  # 行政単独（相続）
    12: 'shiho_touki',     # 司法単独（相続登記）
    13: 'gyosei_yuigon',   # 行政（遺言）
    14: 'rengo_zoyo',      # 行・司連名（贈与・信託）
    15: 'shiho_zoyo',      # 司法（贈与・信託）
    16: 'yuigon_kensaku',  # 遺言検索のみ
}

os.makedirs(OUT_DIR, exist_ok=True)

for idx, key in VARIANTS.items():
    wb = openpyxl.load_workbook(SRC)  # 数式保持
    target = wb.worksheets[idx]
    # 対象シート以外を削除
    for s in list(wb.worksheets):
        if s is not target:
            wb.remove(s)
    ws = wb.worksheets[0]

    # 法定相続情報: 代理人住所(G10)の VLOOKUP を行政住所に解決
    if key == 'houtei':
        ws['G10'] = GYOSEI_ADDR

    # 枠外(Z=26以降)のマージを解除してから値・罫線・塗りを消す（空枠が残らないように）。
    for mr in list(ws.merged_cells.ranges):
        if mr.min_col >= 26:
            ws.unmerge_cells(str(mr))

    # 数式セルと枠外列を除去。コメントも除去（ExcelJSが読めるよう rel 不整合を避ける）。
    for row in ws.iter_rows():
        for c in row:
            if c.column >= 26:
                c.value = None
                c.border = NO_BORDER
                c.fill = NO_FILL
            elif isinstance(c.value, str) and c.value.startswith('='):
                c.value = None
            if c.comment is not None:
                c.comment = None

    out = os.path.join(OUT_DIR, f'{key}.xlsx')
    wb.save(out)
    print(f'  wrote {out}')

print('done', len(VARIANTS), 'files')
