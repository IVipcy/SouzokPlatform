#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Generate 仕様書_独自アプリ版.xlsx - Comprehensive specification for
the inheritance platform using Next.js + Supabase + Claude API + Stripe + Render.
"""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

BASE = r"C:\Users\sugur\Desktop\相続プラットフォーム"
OUT = f"{BASE}\\仕様書_独自アプリ版.xlsx"

# ── Color constants ──
HEADER_FILL = PatternFill(start_color="D5E8F0", end_color="D5E8F0", fill_type="solid")
SECTION_FILL = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, size=10)
BODY_FONT = Font(name="Arial", size=10)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# ── Helper ──
def style_sheet(ws, header_row=1, col_widths=None):
    """Apply standard formatting to a worksheet."""
    # Header
    for cell in ws[header_row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    # Body
    for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = WRAP_ALIGN
            cell.border = THIN_BORDER
    # Column widths
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    # Freeze panes
    ws.freeze_panes = ws.cell(row=header_row+1, column=1).coordinate
    # Auto-filter
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"


def remove_zoho_refs(text):
    """Remove Zoho-specific references from text while keeping business logic."""
    if not isinstance(text, str):
        return text
    # Remove specific Zoho references
    patterns = [
        r'Zoho\s*標準フィールド[。、]?',
        r'Zoho\s*CRM\s*[のでにをは]?',
        r'Zoho\s*標準\s*[のでにをは]?',
        r'Deluge\s*[でにをはの関連]?\s*[自動計算処理生成設定]?',
        r'ブループリント\s*[でにをはの制御遷移]?\s*[自動制御]?',
        r'ワークフロー\s*[でにをはの自動]?\s*[自動設定]?',
        r'Zoho\s*Books?\s*[のでにをは]?',
        r'Zoho\s*Writer\s*[のでにをは]?',
        r'Zoho\s*WorkDrive\s*[のでにをは]?',
        r'Zoho\s*Analytics\s*[のでにをは]?',
        r'Zoho\s*Flow\s*[のでにをは]?',
        r'CustomModule\d+',
        r'CRM\s*[のでにをは]?\s*(?:案件|カレンダー|タスク|テンプレート)',
    ]
    result = text
    for p in patterns:
        result = re.sub(p, '', result)
    # Clean up leftover punctuation
    result = re.sub(r'[。、]{2,}', '。', result)
    result = re.sub(r'^\s*[。、]\s*', '', result)
    result = result.strip()
    return result if result else text


def jp_to_snake(name):
    """Convert Japanese field name to snake_case DB column name."""
    mapping = {
        '案件名': 'case_name',
        '管理番号': 'management_number',
        '案件ステージ': 'case_stage',
        '案件ステータス': 'case_status',
        '受任日': 'engagement_date',
        '完了日': 'completion_date',
        '完了予定日': 'estimated_completion_date',
        '契約形態': 'contract_type',
        '受注担当': 'assigned_staff',
        '管理担当': 'management_staff',
        '受注経路': 'acquisition_channel',
        '紹介元パートナー': 'referral_partner',
        'パートナー企業名': 'partner_company_name',
        'パートナー区分': 'partner_category',
        '担当者名': 'contact_person_name',
        '電話番号': 'phone_number',
        'メールアドレス': 'email',
        '住所': 'address',
        '郵便番号': 'postal_code',
        '紹介手数料率': 'referral_fee_rate',
        '振込先銀行': 'bank_name',
        '支店名': 'branch_name',
        '口座種別': 'account_type',
        '口座番号': 'account_number',
        '口座名義': 'account_holder',
        '被相続人名': 'decedent_name',
        '被相続人ふりがな': 'decedent_name_kana',
        '死亡日': 'death_date',
        '死亡時住所': 'death_address',
        '本籍': 'registered_domicile',
        '筆頭者': 'head_of_family',
        '最終住所': 'last_address',
        '生年月日': 'birth_date',
        '依頼者名': 'client_name',
        '依頼者ふりがな': 'client_name_kana',
        '依頼者続柄': 'client_relationship',
        '依頼者住所': 'client_address',
        '依頼者電話番号': 'client_phone',
        '依頼者メール': 'client_email',
        '相続人数': 'heir_count',
        '法定相続人数': 'legal_heir_count',
        '遺言有無': 'will_exists',
        '遺言種類': 'will_type',
        '遺産分割方法': 'estate_division_method',
        '分割協議ステータス': 'division_agreement_status',
        '相続税申告要否': 'tax_filing_required',
        '申告期限': 'filing_deadline',
        '税理士名': 'tax_accountant_name',
        '不動産有無': 'real_estate_exists',
        '不動産件数': 'real_estate_count',
        '名義変更要否': 'name_change_required',
        '預貯金件数': 'deposit_count',
        '証券有無': 'securities_exists',
        '証券件数': 'securities_count',
        '生命保険有無': 'life_insurance_exists',
        '報酬金額': 'fee_amount',
        '立替実費合計': 'advance_expense_total',
        '確定金額': 'confirmed_amount',
        '入金ステータス': 'payment_status',
        '請求日': 'billing_date',
        '入金日': 'payment_date',
        '入金額': 'payment_amount',
        'パートナー報酬': 'partner_fee',
        'パートナー報酬率': 'partner_fee_rate',
        'パートナー支払ステータス': 'partner_payment_status',
        '使用目的': 'purpose',
        '請求理由': 'request_reason',
        '外字有無': 'special_char_exists',
        '戸籍特記事項': 'koseki_notes',
        '戸籍請求書パターン': 'koseki_request_pattern',
        '請求先市区町村': 'request_municipality',
        '面談日': 'meeting_date',
        '面談種別': 'meeting_type',
        '面談メモ': 'meeting_notes',
        '面談場所': 'meeting_location',
        '初回面談完了': 'initial_meeting_done',
        '委任契約日': 'delegation_contract_date',
        '委任契約書回収': 'delegation_contract_collected',
        '印鑑証明回収': 'seal_certificate_collected',
        '本人確認書類回収': 'id_document_collected',
        'タスク名': 'task_name',
        'タスクカテゴリ': 'task_category',
        'タスクステータス': 'task_status',
        '期限': 'due_date',
        '担当者': 'assignee',
        '優先度': 'priority',
        '手順テキスト': 'procedure_text',
        '作業メモ': 'work_notes',
        '対象タスク': 'target_task',
        '関連案件': 'related_case',
        'フェーズ': 'phase',
        '完了条件': 'completion_criteria',
        '前提タスク': 'prerequisite_task',
        '自動生成': 'auto_generated',
        '相続人氏名': 'heir_name',
        'ふりがな': 'name_kana',
        '続柄': 'relationship',
        '法定相続分': 'legal_share',
        '遺産取得割合': 'estate_share_ratio',
        '金融機関名': 'financial_institution',
        '支店': 'branch',
        '口座種類': 'account_category',
        '残高': 'balance',
        '手続ステータス': 'procedure_status',
        '所在地': 'location',
        '地目': 'land_category',
        '地積': 'land_area',
        '評価額': 'appraised_value',
        '登記名義人': 'registered_owner',
        '備考': 'notes',
        'メモ': 'memo',
        '作成日': 'created_at',
        '更新日': 'updated_at',
        '作成者': 'created_by',
        '更新者': 'updated_by',
        '関連タスク': 'related_task',
        '金額': 'amount',
        '費目': 'expense_category',
        '支払日': 'payment_date_expense',
        '領収書有無': 'receipt_exists',
        '原本受領書回収': 'original_receipt_collected',
        '立替実費計算完了': 'advance_calc_done',
        '請求書作成完了': 'invoice_created',
        '経費区分': 'expense_type',
        '立替日': 'advance_date',
        '立替金額': 'advance_amount',
        '名寄せ請求先': 'nayose_destination',
        '不動産所在地': 'property_location',
        '被相続人との関係': 'relationship_to_decedent',
        '住所（相続人）': 'heir_address',
        '電話番号（相続人）': 'heir_phone',
        '印鑑証明取得済': 'seal_cert_obtained',
        '署名済': 'signed',
        '概要': 'summary',
        '説明': 'description',
        '進捗率': 'progress_rate',
        '開始日': 'start_date',
        '終了日': 'end_date',
        'コメント': 'comments',
    }
    if not isinstance(name, str) or pd.isna(name):
        return ''
    name_clean = name.strip()
    if name_clean in mapping:
        return mapping[name_clean]
    # Fallback: try to create a reasonable snake_case
    # Remove parenthetical content for matching
    base_name = re.sub(r'[（(].*?[）)]', '', name_clean).strip()
    if base_name in mapping:
        return mapping[base_name]
    # Return romanized placeholder
    return ''


def rewrite_impl(text):
    """Rewrite 実装方法 to remove Zoho references and use new tech stack."""
    if not isinstance(text, str) or pd.isna(text):
        return text
    rewrites = {
        '相続ステーションにZohoの案件モジュールへの連携ボタンを設置': '相続ステーションからNext.js APIへWebhook連携',
        'Zohoプラグインの標準機能': 'Next.js + Supabase RLS による案件CRUD',
        'Zoho CRMブループリントで自動遷移': 'Supabase Database Function + Next.js API Route でステージ自動遷移',
        'Zoho CRMブループリント＋Delugeで自動遷移': 'Supabase Database Function + Next.js API Route でステージ自動遷移',
        'ブループリント＋Delugeで自動遷移': 'Supabase Database Function + Next.js API Route でステージ自動遷移',
        'Deluge関数でタスク一括生成': 'Supabase Edge Function でタスク一括生成',
        'タスクモジュール＋ブループリント': 'Supabase タスクテーブル + Next.js API Route',
        'CRMカレンダー＋タスク連動': 'Next.js カレンダーUI + Supabase タスクテーブル連動',
        'CRM関連リストでサブフォーム管理': 'Supabase リレーションテーブル + Next.js 動的フォーム',
        'Zoho Writer差し込み＋テンプレート': 'Claude API による文書テンプレート差し込み生成',
        'Zoho Writer差し込み印刷': 'Claude API による文書テンプレート差し込み生成',
        'Writer差し込み印刷': 'Claude API による文書テンプレート差し込み生成',
        'WorkDriveフォルダ自動作成': 'Supabase Storage + Next.js API Route でフォルダ自動作成',
        'WorkDrive＋共有リンク': 'Supabase Storage + 共有リンク生成',
        'Zoho Books連携': 'Stripe Billing + Next.js API Route',
        'Books請求書＋Deluge自動計算': 'Stripe Billing + Supabase Edge Function で自動計算',
        'Analyticsダッシュボード': 'Next.js ダッシュボード + Supabase集計クエリ',
        'Analytics＋CRMレポート': 'Next.js ダッシュボード + Supabase集計クエリ',
        'Zoho Flowで外部連携': 'Next.js API Route + Webhook で外部連携',
        'RAGシステムにガイドドキュメントを学習させAI回答': 'Claude API + RAGシステム（Supabase pgvector）でAI回答',
        '案件完了時にAPI/Webhookでキントーンへ自動連携': 'Next.js API Route + Webhook でキントーンへ自動連携',
    }
    for old, new in rewrites.items():
        if old in text:
            return new
    # Generic replacements
    result = text
    result = re.sub(r'Zoho\s*(CRM|Books?|Writer|WorkDrive|Analytics|Flow)\s*', '', result)
    result = re.sub(r'Deluge\s*', 'Supabase Edge Function ', result)
    result = re.sub(r'ブループリント', 'ステート管理ロジック', result)
    result = re.sub(r'ワークフロー', 'API Route', result)
    return result.strip()


def get_tech_component(impl_text, func_name=''):
    """Determine 技術コンポーネント based on implementation method."""
    if not isinstance(impl_text, str):
        return ''
    text = str(impl_text) + str(func_name)
    components = []
    if any(k in text for k in ['API Route', 'API', 'Webhook', '連携', 'CRUD']):
        components.append('Next.js API Route')
    if any(k in text for k in ['Supabase', 'RLS', 'テーブル', 'DB', 'Edge Function', 'pgvector']):
        components.append('Supabase')
    if any(k in text for k in ['Claude', 'AI', 'RAG', '文書', 'テンプレート']):
        components.append('Claude API')
    if any(k in text for k in ['Stripe', '請求', '課金', 'Billing']):
        components.append('Stripe Billing')
    if any(k in text for k in ['ダッシュボード', '集計', 'レポート']):
        components.append('Next.js Dashboard')
    if any(k in text for k in ['Storage', 'フォルダ', 'ファイル']):
        components.append('Supabase Storage')
    if any(k in text for k in ['カレンダー']):
        components.append('Next.js Calendar UI')
    if not components:
        components.append('Next.js + Supabase')
    return ' + '.join(components)


# ══════════════════════════════════════════════════════════════════════════════
# READ SOURCE DATA
# ══════════════════════════════════════════════════════════════════════════════
print("Reading source files...")

# 1) 機能別マトリックス
df_func_raw = pd.read_excel(f"{BASE}\\機能別マトリックス (3).xlsx", sheet_name="機能別マトリックス", header=None)

# 2) 画面項目一覧（統合版 - 案件）
df_case_raw = pd.read_excel(f"{BASE}\\画面項目一覧_統合版.xlsx", sheet_name="案件詳細画面（統合版）")

# 3) タスク詳細画面
df_task_raw = pd.read_excel(f"{BASE}\\20260315=2030__画面項目一覧_修正.xlsx", sheet_name="タスク詳細画面", header=None)

# 4) タスク×案件項目マトリクス
df_matrix_raw = pd.read_excel(f"{BASE}\\タスク統合設計書 (1).xlsx", sheet_name="①タスク×案件項目マトリクス", header=None)

# 5) タスク自動生成ロジック
df_logic_raw = pd.read_excel(f"{BASE}\\タスク統合設計書 (1).xlsx", sheet_name="②タスク自動生成ロジック", header=None)

print("Source files read successfully.")

# ══════════════════════════════════════════════════════════════════════════════
# CREATE WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: 機能定義書
# ══════════════════════════════════════════════════════════════════════════════
print("Creating Sheet 1: 機能定義書...")
ws1 = wb.active
ws1.title = "機能定義書"

# Parse the raw data - row 0 is merged header, row 1 is column headers, data from row 2
headers_row = df_func_raw.iloc[0].tolist()
# Columns: #, カテゴリ, 機能名, 機能定義, 解決する課題, 実装方法, 開発スコープ, 優先度, 開発工数, 改善インパクト, then Zoho columns
# Skip row 0 (merged header "機能定義") and row 1 (column header row "#, カテゴリ, ...")
# But from inspection, row 0 in df_func_raw IS the column header row ["#", "カテゴリ", ...]
# because pandas read the first merged cell as "機能定義" but the actual column headers are there
# Let's check: row 0 has "#" in col 0 -> it's the header row. Data starts from row 1.
data_rows = df_func_raw.iloc[1:].reset_index(drop=True)

# But we saw duplicate headers in output row 2. This means df_func_raw[0] = ["機能定義","Unnamed",...]
# and df_func_raw[1] = ["#","カテゴリ",...] which is the real header. Data is from row 2.
# Let's handle: skip any row where col 0 == '#' (header echo)


# Output headers
out_headers_1 = ['#', 'カテゴリ', '機能名', '機能定義', '解決する課題', '実装方法', '技術コンポーネント', '開発スコープ', '優先度', '開発工数', '改善インパクト']
for col_idx, h in enumerate(out_headers_1, 1):
    ws1.cell(row=1, column=col_idx, value=h)

row_out = 2
for i in range(len(data_rows)):
    r = data_rows.iloc[i]
    num = r.iloc[0]
    cat = r.iloc[1]
    fname = r.iloc[2]
    fdef = r.iloc[3]
    issue = r.iloc[4]
    impl_orig = r.iloc[5]
    scope = r.iloc[6]
    pri = r.iloc[7]
    effort = r.iloc[8]
    impact = r.iloc[9]

    if pd.isna(num) and pd.isna(fname):
        continue
    # Skip header echo rows
    if str(num) == '#' or str(num) == '機能定義':
        continue

    impl_new = rewrite_impl(str(impl_orig)) if pd.notna(impl_orig) else ''
    tech = get_tech_component(impl_new, str(fname) if pd.notna(fname) else '')

    # Clean Zoho references from function name and definition
    fname_clean = fname
    if pd.notna(fname):
        fname_clean = str(fname).replace('Zohoの', '').replace('とZohoのシステム連携', 'との自動連携').replace('Zoho', 'システム')
    fdef_clean = fdef
    if pd.notna(fdef):
        fdef_str = str(fdef)
        fdef_str = fdef_str.replace('Zohoに連携し、Zohoで', 'システムに連携し、')
        fdef_str = fdef_str.replace('Zoho', 'システム')
        fdef_clean = fdef_str

    vals = [num, cat, fname_clean, fdef_clean, issue, impl_new, tech, scope, pri, effort, impact]
    for col_idx, v in enumerate(vals, 1):
        ws1.cell(row=row_out, column=col_idx, value=None if (isinstance(v, float) and pd.isna(v)) else v)
    row_out += 1

style_sheet(ws1, col_widths=[5, 14, 35, 50, 40, 45, 35, 12, 8, 15, 25])
print(f"  Sheet 1 done: {row_out-2} rows")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: 画面項目一覧（案件）
# ══════════════════════════════════════════════════════════════════════════════
print("Creating Sheet 2: 画面項目一覧（案件）...")
ws2 = wb.create_sheet("画面項目一覧（案件）")

# Headers: #, セクション, 項目名（日本語）, DB列名, データ型, 選択肢・書式, 必須, 備考・説明
out_headers_2 = ['#', 'セクション', '項目名（日本語）', 'DB列名', 'データ型', '選択肢・書式', '必須', '備考・説明']
for col_idx, h in enumerate(out_headers_2, 1):
    ws2.cell(row=1, column=col_idx, value=h)

row_out = 2
for i in range(len(df_case_raw)):
    r = df_case_raw.iloc[i]
    num = r['#']
    section = r['セクション']
    field_name = r['項目名（日本語）']
    dtype = r['データ型']
    choices = r['選択肢・書式']
    required = r['必須']
    notes = r['備考・説明']

    # Check if this is a section header row (has section-like # like "1. 基本情報")
    is_section_header = False
    if pd.notna(num) and isinstance(num, str) and re.match(r'^\d+\.\s', str(num)):
        is_section_header = True
    if pd.notna(num) and isinstance(num, str) and num.startswith('A.'):
        is_section_header = True
    if pd.notna(num) and isinstance(num, str) and num.startswith('B.'):
        is_section_header = True
    if pd.notna(num) and isinstance(num, str) and num.startswith('C.'):
        is_section_header = True
    if pd.notna(num) and isinstance(num, str) and num.startswith('D.'):
        is_section_header = True
    if pd.notna(num) and isinstance(num, str) and num.startswith('E.'):
        is_section_header = True
    if pd.notna(num) and isinstance(num, str) and num.startswith('F.'):
        is_section_header = True
    if pd.notna(num) and isinstance(num, str) and num.startswith('G.'):
        is_section_header = True
    if pd.notna(num) and isinstance(num, str) and num.startswith('H.'):
        is_section_header = True

    # Skip completely empty rows
    if all(pd.isna(v) for v in [num, section, field_name, dtype]):
        continue

    # Clean data type - replace Zoho-specific types
    dtype_clean = dtype
    if pd.notna(dtype):
        dtype_str = str(dtype)
        dtype_str = dtype_str.replace('ルックアップ（取引先）', 'リレーション（パートナー）')
        dtype_str = dtype_str.replace('ルックアップ（案件）', 'リレーション（案件）')
        dtype_str = dtype_str.replace('ルックアップ（タスク）', 'リレーション（タスク）')
        dtype_str = re.sub(r'ルックアップ', 'リレーション', dtype_str)
        dtype_clean = dtype_str

    # Clean notes
    notes_clean = remove_zoho_refs(notes) if pd.notna(notes) else None

    # DB column name
    db_col = jp_to_snake(field_name) if pd.notna(field_name) else ''

    vals = [
        None if (isinstance(num, float) and pd.isna(num)) else num,
        None if pd.isna(section) else section,
        None if pd.isna(field_name) else field_name,
        db_col if db_col else None,
        None if pd.isna(dtype_clean) else dtype_clean,
        None if pd.isna(choices) else choices,
        None if pd.isna(required) else required,
        notes_clean,
    ]
    for col_idx, v in enumerate(vals, 1):
        ws2.cell(row=row_out, column=col_idx, value=v)

    # Apply section header formatting
    if is_section_header:
        for col_idx in range(1, len(out_headers_2) + 1):
            cell = ws2.cell(row=row_out, column=col_idx)
            cell.fill = SECTION_FILL
            cell.font = Font(name="Arial", bold=True, size=10)

    row_out += 1

style_sheet(ws2, col_widths=[8, 18, 25, 30, 22, 35, 6, 50])
print(f"  Sheet 2 done: {row_out-2} rows")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: 画面項目一覧（タスク）
# ══════════════════════════════════════════════════════════════════════════════
print("Creating Sheet 3: 画面項目一覧（タスク）...")
ws3 = wb.create_sheet("画面項目一覧（タスク）")

# The task sheet has a different structure - header is in row 1 (0-indexed)
# Columns from inspection: #, セクション, 項目名（日本語）, データ型, 選択肢・書式, 必須, 備考・説明, 出典, 対象タスク
out_headers_3 = ['#', 'セクション', '項目名（日本語）', 'DB列名', 'データ型', '選択肢・書式', '必須', '備考・説明', '対象タスク']
for col_idx, h in enumerate(out_headers_3, 1):
    ws3.cell(row=1, column=col_idx, value=h)

row_out = 2
for i in range(len(df_task_raw)):
    r = df_task_raw.iloc[i]
    # Row 0 is the title row, row 1 has actual headers
    if i <= 1:
        continue  # skip title and header rows

    num = r.iloc[0]
    section = r.iloc[1]
    field_name = r.iloc[2]
    dtype = r.iloc[3]
    choices = r.iloc[4]
    required = r.iloc[5]
    notes = r.iloc[6]
    # source = r.iloc[7]  # 出典 - skip
    target_task = r.iloc[8]

    # Skip header echo rows
    if pd.notna(num) and str(num) == '#':
        continue

    # Check section header
    is_section_header = False
    if pd.notna(num) and isinstance(num, str) and re.match(r'^\d+\.\s', str(num)):
        is_section_header = True

    # Skip empty rows
    if all(pd.isna(v) for v in [num, section, field_name, dtype]):
        continue

    # Clean dtype
    dtype_clean = dtype
    if pd.notna(dtype):
        dtype_str = str(dtype)
        dtype_str = dtype_str.replace('ルックアップ（取引先）', 'リレーション（パートナー）')
        dtype_str = dtype_str.replace('ルックアップ（案件）', 'リレーション（案件）')
        dtype_str = dtype_str.replace('ルックアップ（タスク）', 'リレーション（タスク）')
        dtype_str = re.sub(r'ルックアップ', 'リレーション', dtype_str)
        dtype_clean = dtype_str

    notes_clean = remove_zoho_refs(notes) if pd.notna(notes) else None
    db_col = jp_to_snake(field_name) if pd.notna(field_name) else ''

    vals = [
        None if (isinstance(num, float) and pd.isna(num)) else num,
        None if pd.isna(section) else section,
        None if pd.isna(field_name) else field_name,
        db_col if db_col else None,
        None if pd.isna(dtype_clean) else dtype_clean,
        None if pd.isna(choices) else choices,
        None if pd.isna(required) else required,
        notes_clean,
        None if pd.isna(target_task) else target_task,
    ]
    for col_idx, v in enumerate(vals, 1):
        ws3.cell(row=row_out, column=col_idx, value=v)

    if is_section_header:
        for col_idx in range(1, len(out_headers_3) + 1):
            cell = ws3.cell(row=row_out, column=col_idx)
            cell.fill = SECTION_FILL
            cell.font = Font(name="Arial", bold=True, size=10)

    row_out += 1

style_sheet(ws3, col_widths=[8, 15, 25, 30, 22, 35, 6, 50, 18])
print(f"  Sheet 3 done: {row_out-2} rows")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4: タスク自動生成ロジック
# ══════════════════════════════════════════════════════════════════════════════
print("Creating Sheet 4: タスク自動生成ロジック...")
ws4 = wb.create_sheet("タスク自動生成ロジック")

out_headers_4 = ['#', 'フェーズ', 'タスク名', '案件から参照する項目', '生成条件', '生成数', '前提タスク', '備考']
for col_idx, h in enumerate(out_headers_4, 1):
    ws4.cell(row=1, column=col_idx, value=h)

# Module name replacements
module_replacements = {
    'CustomModule1': '相続人テーブル',
    'CustomModule2': '預貯金テーブル',
    'CustomModule3': '証券テーブル',
    'CustomModule4': '生命保険テーブル',
    'CustomModule5': '不動産テーブル',
    'CustomModule6': '立替実費テーブル',
    'CustomModule7': '分割内容テーブル',
}

def clean_module_refs(text):
    if not isinstance(text, str) or pd.isna(text):
        return text
    result = text
    for old, new in module_replacements.items():
        result = result.replace(old, new)
    result = remove_zoho_refs(result)
    return result

# Build merged data from matrix + logic sheets
# Matrix: #, フェーズ, タスク名, 案件詳細から参照する項目, モジュールから参照する項目, 参照先モジュール, 備考
# Logic: #, フェーズ, タスク名, 生成条件, 生成数, 前提タスク, 備考

# Parse matrix (rows 0=title, 1=empty, 2=headers -> data starts at 3)
# But some sheets have headers at row 1 or 2. Skip any row where col 0 is '#' or NaN+NaN
matrix_data = {}
for i in range(2, len(df_matrix_raw)):
    r = df_matrix_raw.iloc[i]
    num = r.iloc[0]
    phase = r.iloc[1]
    tname = r.iloc[2]
    ref_items = r.iloc[3]  # 案件詳細から参照する項目
    mod_ref = r.iloc[4]  # モジュールから参照する項目
    ref_module = r.iloc[5]  # 参照先モジュール

    if pd.isna(num) and pd.isna(tname):
        continue
    # Skip header rows
    if pd.notna(num) and str(num).strip() == '#':
        continue
    # Combine ref items
    combined_ref = ''
    if pd.notna(ref_items):
        combined_ref = str(ref_items)
    if pd.notna(mod_ref) and str(mod_ref).strip() not in ['なし', 'ー', '-', '－']:
        if combined_ref:
            combined_ref += '\n'
        module_name = clean_module_refs(str(ref_module)) if pd.notna(ref_module) else ''
        combined_ref += f"[{module_name}] {clean_module_refs(str(mod_ref))}"

    key = str(tname).strip() if pd.notna(tname) else str(num)
    matrix_data[key] = {
        'num': num,
        'phase': phase,
        'task_name': tname,
        'ref_items': combined_ref,
    }

# Parse logic
logic_data = {}
for i in range(2, len(df_logic_raw)):
    r = df_logic_raw.iloc[i]
    num = r.iloc[0]
    phase = r.iloc[1]
    tname = r.iloc[2]
    condition = r.iloc[3]
    count = r.iloc[4]
    prereq = r.iloc[5]
    note = r.iloc[6]

    if pd.isna(num) and pd.isna(tname):
        continue
    # Skip header rows
    if pd.notna(num) and str(num).strip() == '#':
        continue

    key = str(tname).strip() if pd.notna(tname) else str(num)
    logic_data[key] = {
        'num': num,
        'phase': phase,
        'task_name': tname,
        'condition': condition,
        'count': count,
        'prereq': prereq,
        'note': note,
    }

# Merge and write - use logic as the primary, supplement with matrix
row_out = 2
phase_headers_written = set()

for i in range(2, len(df_logic_raw)):
    r = df_logic_raw.iloc[i]
    num = r.iloc[0]
    phase = r.iloc[1]
    tname = r.iloc[2]
    condition = r.iloc[3]
    count = r.iloc[4]
    prereq = r.iloc[5]
    note = r.iloc[6]

    # Skip header echo rows
    if pd.notna(num) and str(num).strip() == '#':
        continue

    # Phase header row
    is_phase_header = False
    if pd.notna(num) and isinstance(num, str) and 'Phase' in str(num):
        is_phase_header = True
    if pd.isna(num) and pd.isna(tname) and pd.notna(phase):
        is_phase_header = True

    if is_phase_header:
        phase_text = str(num) if pd.notna(num) else str(phase)
        phase_text = clean_module_refs(phase_text)
        ws4.cell(row=row_out, column=1, value=phase_text)
        for col_idx in range(1, len(out_headers_4) + 1):
            cell = ws4.cell(row=row_out, column=col_idx)
            cell.fill = SECTION_FILL
            cell.font = Font(name="Arial", bold=True, size=10)
        row_out += 1
        continue

    if pd.isna(num) and pd.isna(tname):
        continue

    # Get matrix info for this task
    key = str(tname).strip() if pd.notna(tname) else ''
    m_info = matrix_data.get(key, {})

    ref_items = m_info.get('ref_items', '')
    ref_items = clean_module_refs(ref_items) if ref_items else ''

    condition_clean = clean_module_refs(condition) if pd.notna(condition) else None
    note_clean = clean_module_refs(note) if pd.notna(note) else None

    vals = [
        None if (isinstance(num, float) and pd.isna(num)) else num,
        None if pd.isna(phase) else phase,
        None if pd.isna(tname) else tname,
        ref_items if ref_items else None,
        condition_clean,
        None if pd.isna(count) else count,
        None if pd.isna(prereq) else prereq,
        note_clean,
    ]
    for col_idx, v in enumerate(vals, 1):
        ws4.cell(row=row_out, column=col_idx, value=v)
    row_out += 1

style_sheet(ws4, col_widths=[5, 12, 25, 45, 35, 10, 20, 40])
print(f"  Sheet 4 done: {row_out-2} rows")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5: 画面遷移図
# ══════════════════════════════════════════════════════════════════════════════
print("Creating Sheet 5: 画面遷移図...")
ws5 = wb.create_sheet("画面遷移図")

out_headers_5 = ['#', '遷移元画面', '操作/トリガー', '遷移先画面', '備考']
for col_idx, h in enumerate(out_headers_5, 1):
    ws5.cell(row=1, column=col_idx, value=h)

transitions = [
    # Login
    (1, 'login', 'ログイン成功', 'dashboard', 'Supabase Auth認証後リダイレクト'),
    (2, 'login', 'パスワードリセット', 'password_reset', 'メールリンクによるリセット'),
    # Dashboard
    (3, 'dashboard', '案件カードをクリック', 'case_detail', '案件詳細画面へ遷移'),
    (4, 'dashboard', '「新規面談」ボタン', 'intake', '新規面談受付フォーム'),
    (5, 'dashboard', 'サイドバー「スケジュール」', 'schedule', 'カレンダー表示'),
    (6, 'dashboard', 'サイドバー「タスク」', 'task_list', 'タスク一覧画面'),
    (7, 'dashboard', 'サイドバー「請求・売上」', 'billing', 'Stripe連携請求管理'),
    (8, 'dashboard', 'サイドバー「レポート」', 'report', '集計ダッシュボード'),
    (9, 'dashboard', 'サイドバー「パートナー」', 'partner_list', 'パートナー一覧'),
    (10, 'dashboard', 'サイドバー「ナレッジAI」', 'knowledge_ai', 'AIチャット画面'),
    # Case detail
    (11, 'case_detail', 'タスクタブ', 'case_detail#tasks', 'タスク一覧（案件内）'),
    (12, 'case_detail', '相続人タブ', 'case_detail#heirs', '相続人サブテーブル編集'),
    (13, 'case_detail', '財産タブ', 'case_detail#assets', '預貯金・不動産・証券'),
    (14, 'case_detail', '書類生成ボタン', 'document_generate', 'Claude APIで文書差し込み生成'),
    (15, 'case_detail', '請求書作成ボタン', 'billing_create', 'Stripe請求書作成'),
    (16, 'case_detail', 'タスク自動生成ボタン', 'case_detail', 'Edge Functionでタスク一括生成'),
    (17, 'case_detail', '戻るボタン', 'dashboard', 'ダッシュボードへ戻る'),
    # Intake
    (18, 'intake', '面談情報入力→保存', 'case_detail', '新規案件レコード作成後に遷移'),
    (19, 'intake', 'キャンセル', 'dashboard', '入力内容破棄'),
    # Task list
    (20, 'task_list', 'タスク行をクリック', 'task_detail', 'タスク詳細画面'),
    (21, 'task_list', 'フィルタ変更', 'task_list', '担当者・ステータス・フェーズで絞り込み'),
    # Task detail
    (22, 'task_detail', 'ステータス変更', 'task_detail', '完了→次タスク自動生成トリガー'),
    (23, 'task_detail', '関連案件リンク', 'case_detail', '紐づく案件へ遷移'),
    (24, 'task_detail', '戻る', 'task_list', 'タスク一覧へ戻る'),
    # Schedule
    (25, 'schedule', '予定をクリック', 'task_detail', '関連タスク詳細へ遷移'),
    (26, 'schedule', '日付をクリック', 'schedule#new', '新規予定登録モーダル'),
    # Billing
    (27, 'billing', '請求書詳細', 'billing_detail', 'Stripe請求書詳細'),
    (28, 'billing', '新規請求作成', 'billing_create', 'Stripe Invoice作成フォーム'),
    # Report
    (29, 'report', 'ドリルダウン', 'case_detail', '集計数値→対象案件一覧→詳細'),
    # Partner
    (30, 'partner_list', 'パートナー行クリック', 'partner_detail', 'パートナー詳細画面'),
    (31, 'partner_detail', '関連案件タブ', 'partner_detail#cases', '紐づく案件一覧'),
    # Knowledge AI
    (32, 'knowledge_ai', '質問送信', 'knowledge_ai', 'Claude API + pgvectorでRAG回答'),
    # Common sidebar navigation
    (33, '全画面（共通）', 'サイドバー「ダッシュボード」', 'dashboard', '共通ナビゲーション'),
    (34, '全画面（共通）', 'サイドバー「スケジュール」', 'schedule', '共通ナビゲーション'),
    (35, '全画面（共通）', 'サイドバー「タスク」', 'task_list', '共通ナビゲーション'),
    (36, '全画面（共通）', 'サイドバー「新規面談」', 'intake', '共通ナビゲーション'),
    (37, '全画面（共通）', 'サイドバー「請求・売上」', 'billing', '共通ナビゲーション'),
    (38, '全画面（共通）', 'サイドバー「レポート」', 'report', '共通ナビゲーション'),
    (39, '全画面（共通）', 'サイドバー「パートナー」', 'partner_list', '共通ナビゲーション'),
    (40, '全画面（共通）', 'サイドバー「ナレッジAI」', 'knowledge_ai', '共通ナビゲーション'),
    (41, '全画面（共通）', 'ヘッダー「ログアウト」', 'login', 'セッション破棄'),
]

for t in transitions:
    for col_idx, v in enumerate(t, 1):
        ws5.cell(row=t[0]+1, column=col_idx, value=v)

style_sheet(ws5, col_widths=[5, 22, 30, 25, 45])
print(f"  Sheet 5 done: {len(transitions)} rows")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6: 技術スタック
# ══════════════════════════════════════════════════════════════════════════════
print("Creating Sheet 6: 技術スタック...")
ws6 = wb.create_sheet("技術スタック")

out_headers_6 = ['技術', '役割', '採用理由', '主な利用箇所']
for col_idx, h in enumerate(out_headers_6, 1):
    ws6.cell(row=1, column=col_idx, value=h)

tech_data = [
    ('Next.js (App Router)', 'フロントエンド + API', 'Reactベースで1つのフレームワークでUI・APIルート・SSR/SSGを統合開発。TypeScript対応でコード品質を担保。', '全画面UI（ダッシュボード・案件詳細・タスク管理等）、APIルート（/api/*）、ミドルウェア認証チェック'),
    ('PostgreSQL / Supabase', 'データベース + 認証 + ストレージ', 'マルチテナント対応のRLS（Row Level Security）で事務所間データ分離。Supabase Authで認証・SSO対応。Edge Functionでサーバーレス処理。', 'データ永続化（案件・タスク・相続人・財産等全テーブル）、RLSによるアクセス制御、Supabase Auth（ログイン・セッション管理）、Supabase Storage（書類ファイル保管）'),
    ('Claude API (Anthropic)', 'AI機能', '高精度な日本語理解と生成能力。法的文書の差し込みテンプレート生成、業務ナレッジのRAG検索・回答に最適。', '文書自動生成（戸籍請求書・委任契約書・遺産分割協議書等のテンプレート差し込み）、ナレッジAIガイド（銀行手続きガイド等のRAG検索）'),
    ('Stripe', '決済・サブスクリプション管理', '月額課金の自動処理、請求書発行、サブスクリプション管理を外部SaaSで安全に処理。PCI DSS準拠。', '事務所向け月額サブスクリプション課金、請求書発行・管理、売上レポート連携'),
    ('Render', 'ホスティング・インフラ', '500事務所規模に対応可能なスケーラビリティ。Docker対応で低コスト運用。自動デプロイ（GitHub連携）。', 'Next.jsアプリケーションのデプロイ、本番/ステージング環境管理、自動スケーリング'),
    ('Supabase Edge Functions', 'サーバーレス処理', 'タスク自動生成やステージ遷移など非同期バッチ処理をサーバーレスで実行。Denoランタイム。', 'タスク一括自動生成、ステージ自動遷移ロジック、Webhook受信処理、外部API連携（キントーン等）'),
    ('pgvector (Supabase拡張)', 'ベクトル検索', 'PostgreSQL上でベクトル類似度検索を実現。RAGシステムのEmbedding格納・検索に使用。', 'ナレッジAIガイドのドキュメントEmbedding格納、類似度検索によるコンテキスト取得'),
    ('TypeScript', '開発言語', 'フロントエンド・バックエンド共通の型安全な開発。バグ低減とコード補完による開発効率向上。', 'Next.jsアプリ全体（UI・API・型定義）'),
]

for i, t in enumerate(tech_data, 2):
    for col_idx, v in enumerate(t, 1):
        ws6.cell(row=i, column=col_idx, value=v)

style_sheet(ws6, col_widths=[28, 28, 55, 65])
print(f"  Sheet 6 done: {len(tech_data)} rows")

# ══════════════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════════════
wb.save(OUT)
print(f"\nSaved: {OUT}")
print("Done!")
