import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

f = 'C:/Users/sugur/Desktop/相続プラットフォーム/20260315=2030__画面項目一覧_修正.xlsx'
df_main = pd.read_excel(f, sheet_name='案件詳細画面', header=None)
df_sub = pd.read_excel(f, sheet_name='関連モジュール（子レコード）', header=None)

# Parse sub-groups A-H
sub_groups = {}
current_group = None
for i, row in df_sub.iterrows():
    v0 = str(row[0]) if pd.notna(row[0]) else ''
    if i == 0:
        continue
    for letter in ['A','B','C','D','E','F','G','H']:
        if v0.startswith(f'{letter}.'):
            current_group = letter
            sub_groups[letter] = {'title': v0, 'rows': []}
            break
    else:
        if current_group and pd.notna(row[0]):
            try:
                int(row[0])
                sub_groups[current_group]['rows'].append(row)
            except (ValueError, TypeError):
                pass

def make_sub_rows(grp_key):
    if grp_key not in sub_groups:
        return []
    grp = sub_groups[grp_key]
    title = grp['title'].split('\u203b')[0].strip() if '\u203b' in grp['title'] else grp['title'].strip()
    rows = []
    rows.append([f'  \u2514 {title}', None, None, None, None, None, '\u203b\u6848\u4ef6\u306b\u7d10\u3065\u304f\u5b50\u30c6\u30fc\u30d6\u30eb', None, None])
    for sr in grp['rows']:
        rows.append([
            sr[0], sr[1], sr[2], sr[3],
            sr[4] if pd.notna(sr[4]) else None,
            None,
            sr[5] if pd.notna(sr[5]) else None,
            sr[6] if pd.notna(sr[6]) else None,
            sr[7] if pd.notna(sr[7]) else None
        ])
    rows.append([None]*9)
    return rows

# Build merged data
merged_rows = []
merged_rows.append(df_main.iloc[0].tolist())

current_section = 0
inserted = set()

for i in range(1, len(df_main)):
    row = df_main.iloc[i]
    v = str(row[0]) if pd.notna(row[0]) else ''

    new_section = None
    for s in range(1, 18):
        if v.startswith(f'{s}. '):
            new_section = s
            break

    if new_section is not None:
        # Insert sub-groups for the PREVIOUS section before starting new one
        if current_section == 4 and 'A' not in inserted:
            merged_rows.extend(make_sub_rows('A'))
            inserted.add('A')
        elif current_section == 10 and 'E' not in inserted:
            merged_rows.extend(make_sub_rows('E'))
            inserted.add('E')
        elif current_section == 11 and 'BCD' not in inserted:
            for g in ['B', 'C', 'D']:
                merged_rows.extend(make_sub_rows(g))
            inserted.add('BCD')
        elif current_section == 12 and 'F' not in inserted:
            merged_rows.extend(make_sub_rows('F'))
            inserted.add('F')
        current_section = new_section

    merged_rows.append(row.tolist())

# After last row, add G and H
for g in ['G', 'H']:
    merged_rows.extend(make_sub_rows(g))

# Create Excel
wb = Workbook()
ws = wb.active
ws.title = '\u6848\u4ef6\u8a73\u7d30\u753b\u9762\uff08\u7d71\u5408\u7248\uff09'

header_font = Font(bold=True, size=11, name='Arial')
header_fill = PatternFill('solid', fgColor='D5E8F0')
section_font = Font(bold=True, size=11, name='Arial', color='1A5276')
section_fill = PatternFill('solid', fgColor='EBF5FB')
sub_header_font = Font(bold=True, size=10, name='Arial', color='7D3C98')
sub_header_fill = PatternFill('solid', fgColor='F4ECF7')
normal_font = Font(size=10, name='Arial')
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

col_widths = [8, 16, 22, 18, 40, 5, 45, 14, 20]
for idx, w in enumerate(col_widths):
    ws.column_dimensions[get_column_letter(idx+1)].width = w

for r_idx, row_data in enumerate(merged_rows):
    for c_idx, val in enumerate(row_data):
        cell = ws.cell(row=r_idx+1, column=c_idx+1)
        if isinstance(val, float) and pd.isna(val):
            cell.value = None
        elif val is None:
            cell.value = None
        else:
            cell.value = val
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    if r_idx == 0:
        for c_idx in range(9):
            c = ws.cell(row=1, column=c_idx+1)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center', vertical='center')
        continue

    v0 = str(row_data[0]) if row_data[0] is not None and not (isinstance(row_data[0], float) and pd.isna(row_data[0])) else ''

    is_section = False
    for s in range(1, 18):
        if v0.startswith(f'{s}. '):
            is_section = True
            break
    if is_section:
        for c_idx in range(9):
            c = ws.cell(row=r_idx+1, column=c_idx+1)
            c.font = section_font
            c.fill = section_fill

    if '\u2514' in v0:
        for c_idx in range(9):
            c = ws.cell(row=r_idx+1, column=c_idx+1)
            c.font = sub_header_font
            c.fill = sub_header_fill

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:I{len(merged_rows)}'

# Copy other sheets
orig_wb = load_workbook(f)
for sname in orig_wb.sheetnames:
    if sname not in ['\u6848\u4ef6\u8a73\u7d30\u753b\u9762', '\u95a2\u9023\u30e2\u30b8\u30e5\u30fc\u30eb\uff08\u5b50\u30ec\u30b3\u30fc\u30c9\uff09']:
        src = orig_wb[sname]
        dst = wb.create_sheet(sname)
        for row in src.iter_rows():
            for cell in row:
                dst.cell(row=cell.row, column=cell.column, value=cell.value)

out = 'C:/Users/sugur/Desktop/相続プラットフォーム/画面項目一覧_統合版.xlsx'
wb.save(out)
print(f'\u4f5c\u6210\u5b8c\u4e86: {out}')
print(f'\u7d71\u5408\u30b7\u30fc\u30c8: {len(merged_rows)}\u884c')

for g in ['A','B','C','D','E','F','G','H']:
    if g in sub_groups:
        print(f'  {g}. {sub_groups[g]["title"][:30]}... ({len(sub_groups[g]["rows"])}\u9805\u76ee)')
