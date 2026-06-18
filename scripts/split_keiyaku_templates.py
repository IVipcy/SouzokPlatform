# -*- coding: utf-8 -*-
"""
契約書テンプレ分割スクリプト（一回限り）

委任契約書シートを変種ごとに public/templates/keiyaku/<key>.xlsx として切り出す。
契約書は2段組（左A〜Y／右AB〜BB）で右側も本文のため、枠外（参照マスタ・外字ヘルパー）は
BC列(=col55)以降と判定して除去する。数式・コメント・案件番号の見本も除去する。

単独遺言(6)のみ 受任者署名(AD40/AD41)が =IF(BC10=...) で行政/司法を切替える数式のため、
行政書士の署名リテラルに解決してから数式を消す。
"""
import openpyxl
from openpyxl.styles import Border, PatternFill
import os

NO_BORDER = Border()
NO_FILL = PatternFill(fill_type=None)

SRC = os.path.join('docs', '作成書類一覧（契約書・委任状・請求書・領収書）.xlsx')
OUT_DIR = os.path.join('public', 'templates', 'keiyaku')
GYOSEI_SIG = '行政書士法人オーシャン　　　代表社員　　黒田　美菜子'
GYOSEI_AFFIL = '神奈川県行政書士会所属'

# index -> (key, 案件番号見本セル, 枠外閾値col)
VARIANTS = {
    3: ('rengo_zaicho_ari',   ['B58', 'C58', 'D58', 'E58', 'F58'], 55),
    4: ('rengo_zaicho_nashi', ['B52', 'C52', 'D52', 'E52', 'F52'], 55),
    5: ('gyosei_zaicho_ari',  ['B55', 'C55', 'D55', 'E55', 'F55'], 55),
    6: ('tanpoku_yuigon',     ['B47', 'C47', 'D47', 'E47', 'F47'], 55),
}

os.makedirs(OUT_DIR, exist_ok=True)

for idx, (key, strip_cells, off_col) in VARIANTS.items():
    wb = openpyxl.load_workbook(SRC)
    target = wb.worksheets[idx]
    for s in list(wb.worksheets):
        if s is not target:
            wb.remove(s)
    ws = wb.worksheets[0]

    # 単独遺言: 受任者署名の IF を行政書士リテラルに解決
    if key == 'tanpoku_yuigon':
        ws['AD40'] = GYOSEI_SIG
        ws['AD41'] = GYOSEI_AFFIL

    # 枠外列のマージを解除してから値・罫線・塗りを消す（空枠が残らないように）。
    for mr in list(ws.merged_cells.ranges):
        if mr.min_col >= off_col:
            ws.unmerge_cells(str(mr))

    # 数式・コメント・枠外列を除去。
    for row in ws.iter_rows():
        for c in row:
            if c.column >= off_col:
                c.value = None
                c.border = NO_BORDER
                c.fill = NO_FILL
            elif isinstance(c.value, str) and c.value.startswith('='):
                c.value = None
            if c.comment is not None:
                c.comment = None

    # 案件番号の見本（R / - 等のリテラル）を除去
    for addr in strip_cells:
        ws[addr] = None

    out = os.path.join(OUT_DIR, f'{key}.xlsx')
    wb.save(out)
    print(f'  wrote {out}')

print('done', len(VARIANTS), 'files')
